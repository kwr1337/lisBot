from fastapi import APIRouter, Request, Form, HTTPException, Query, File, UploadFile
from fastapi.responses import RedirectResponse, JSONResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from database.models import get_db, log_admin_action
from starlette.middleware.sessions import SessionMiddleware
import logging
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.comments import Comment
from io import BytesIO

router = APIRouter()
templates = Jinja2Templates(directory="./admin_panel/templates")

def is_admin(request: Request) -> bool:
    return request.session.get('is_admin', False)

def get_admin_info(request: Request) -> dict:
    return {
        'username': request.session.get('username', 'Admin'),
        'user_id': request.session.get('user_id')
    }

# Остальные маршруты остаются теми же... 

@router.get("/suggestions")
async def suggestions_page(
    request: Request,
    status: str = Query(default=None),
    date_from: str = Query(default=None),
    date_to: str = Query(default=None)
):
    if not is_admin(request):
        return RedirectResponse(url="/login")
    
    admin_info = get_admin_info(request)
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        query = """
            SELECT 
                bs.id,
                bs.title,
                bs.url,
                bs.price,
                bs.reason,
                bs.status,
                strftime('%d.%m.%Y', bs.created_at) as created_at,
                u.full_name,
                u.username
            FROM book_suggestions bs
            JOIN users u ON bs.user_id = u.id
            WHERE 1=1
        """
        params = []

        # Добавляем отладочные логи
        logging.debug(f"Status: {status}, date_from: {date_from}, date_to: {date_to}")
        
        if status and status != 'all':
            query += " AND bs.status = ?"
            params.append(status)
            logging.debug(f"Added status filter: {status}")
            
        if date_from:
            query += " AND date(bs.created_at) >= date(?)"
            params.append(date_from)
            logging.debug(f"Added date_from filter: {date_from}")
            
        if date_to:
            query += " AND date(bs.created_at) <= date(?)"
            params.append(date_to)
            logging.debug(f"Added date_to filter: {date_to}")
        
        query += " ORDER BY bs.created_at DESC"
        
        logging.debug(f"Final query: {query}")
        logging.debug(f"Params: {params}")
        
        cursor.execute(query, params)
        
        suggestions = []
        for row in cursor.fetchall():
            suggestions.append({
                'id': row[0],
                'title': row[1],
                'url': row[2],
                'price': row[3],
                'reason': row[4],
                'status': row[5],
                'created_at': row[6],
                'user_name': row[7] or f"@{row[8]}"
            })
            
        return templates.TemplateResponse(
            "suggestions.html",
            {
                "request": request,
                "suggestions": suggestions,
                "admin_info": admin_info,
                "status": status or '',
                "date_from": date_from or '',
                "date_to": date_to or ''
            }
        )
    finally:
        conn.close() 

@router.get("/books")
async def books_page(request: Request):
    if not is_admin(request):
        return RedirectResponse(url="/login")
        
    admin_info = get_admin_info(request)
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            SELECT 
                b.id,
                b.title,
                b.author,
                b.theme,
                b.description,
                bc.id as copy_id,
                COALESCE(bb.status, bc.status) as status,
                bb.user_id,
                u.username,
                u.full_name
            FROM books b
            JOIN book_copies bc ON b.id = bc.book_id
            LEFT JOIN (
                SELECT * FROM borrowed_books 
                WHERE status IN ('borrowed', 'booked')
                AND id IN (
                    SELECT MAX(id) 
                    FROM borrowed_books 
                    GROUP BY copy_id
                )
            ) bb ON bc.id = bb.copy_id
            LEFT JOIN users u ON bb.user_id = u.id
            ORDER BY b.title
        """)
        
        books = []
        for row in cursor.fetchall():
            book = {
                'id': row[0],
                'title': row[1],
                'author': row[2],
                'theme': row[3],
                'description': row[4],
                'copy_id': row[5],
                'status': row[6],
                'user': {
                    'id': row[7],
                    'username': row[8],
                    'full_name': row[9]
                } if row[7] else None
            }
            books.append(book)
            
        return templates.TemplateResponse(
            "books.html",
            {
                "request": request,
                "books": books,
                "admin_info": admin_info
            }
        )
    finally:
        conn.close() 

@router.get("/books/{book_id}/qrcodes")
async def book_qrcodes(request: Request, book_id: int):
    if not is_admin(request):
        return RedirectResponse(url="/login")
        
    admin_info = get_admin_info(request)
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            WITH LastBookStatus AS (
                SELECT 
                    copy_id,
                    status,
                    user_id
                FROM borrowed_books
                WHERE id IN (
                    SELECT MAX(id)
                    FROM borrowed_books
                    GROUP BY copy_id
                )
            )
            SELECT 
                b.title,
                b.author,
                bc.id as copy_id,
                COALESCE(lbs.status, bc.status) as status,
                u.id as user_id,
                u.username,
                u.full_name
            FROM books b
            JOIN book_copies bc ON b.id = bc.book_id
            LEFT JOIN LastBookStatus lbs ON bc.id = lbs.copy_id
            LEFT JOIN users u ON lbs.user_id = u.id
            WHERE b.id = ?
            ORDER BY bc.id
        """, (book_id,))
        
        book_copies = []
        book_title = None
        book_author = None
        
        for row in cursor.fetchall():
            if not book_title:
                book_title = row[0]
                book_author = row[1]
                
            copy_info = {
                'id': row[2],
                'status': row[3],
                'user': {
                    'id': row[4],
                    'username': row[5],
                    'full_name': row[6]
                } if row[4] else None
            }
            book_copies.append(copy_info)
            
        return templates.TemplateResponse(
            "book_qrcodes.html",
            {
                "request": request,
                "book_id": book_id,
                "title": book_title,
                "author": book_author,
                "copies": book_copies,
                "admin_info": admin_info
            }
        )
    finally:
        conn.close() 

@router.get("/users")
async def users_page(request: Request):
    if not is_admin(request):
        return RedirectResponse(url="/login")
        
    admin_info = get_admin_info(request)
    
    with get_db() as conn:
        cursor = conn.cursor()
        
        # Отладочный запрос - проверяем все роли и количество пользователей с каждой ролью
        cursor.execute("""
            SELECT role, COUNT(*) as count 
            FROM users 
            GROUP BY role
        """)
        role_debug = cursor.fetchall()
        print("Роли в базе:", role_debug)  # Для отладки в консоли
        
        # Получаем статистику по ролям
        cursor.execute("""
            SELECT 
                (SELECT COUNT(*) FROM users) as total,
                (SELECT COUNT(*) FROM users WHERE role = 'admin') as admins,
                (SELECT COUNT(*) FROM users WHERE role = 'teacher') as teachers,
                (SELECT COUNT(*) FROM users WHERE role = 'user' OR role IS NULL OR role = '') as students
        """)
        stats = cursor.fetchone()
        
        # Получаем список пользователей
        cursor.execute("""
            SELECT 
                u.id,
                u.username,
                u.full_name,
                u.phone,
                u.role,
                (SELECT COUNT(*) FROM borrowed_books WHERE user_id = u.id AND status = 'borrowed') as books_count,
                (SELECT COUNT(*) FROM book_reviews WHERE user_id = u.id) as reviews_count,
                COALESCE(MAX(bb.borrow_date), 'Нет активности') as last_activity,
                0 as is_blocked
            FROM users u
            LEFT JOIN borrowed_books bb ON u.id = bb.user_id
            GROUP BY u.id
            ORDER BY 
                CASE u.role
                    WHEN 'admin' THEN 1
                    WHEN 'teacher' THEN 2
                    ELSE 3
                END,
                u.full_name
        """)
        users = cursor.fetchall()
        
        return templates.TemplateResponse(
            "users.html",
            {
                "request": request,
                "users": users,
                "admin_info": admin_info,
                "stats": stats
            }
        ) 

@router.post("/users/{user_id}/toggle_block")
async def toggle_user_block(request: Request, user_id: int, reason: str = Form(...)):
    if not is_admin(request):
        return RedirectResponse(url="/login")
    
    with get_db() as conn:
        cursor = conn.cursor()
        
        # Получаем текущий статус блокировки
        cursor.execute("SELECT is_blocked FROM users WHERE id = ?", (user_id,))
        user = cursor.fetchone()
        
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        
        is_blocked = user[0]
        new_status = 0 if is_blocked else 1
        
        # Обновляем статус блокировки
        cursor.execute("UPDATE users SET is_blocked = ? WHERE id = ?", (new_status, user_id))
        conn.commit()
        
        # Отправляем уведомление пользователю
        try:
            bot = request.app.state.bot  # Получаем экземпляр бота из состояния приложения
            if new_status == 1:
                await bot.send_message(user_id, f"🚫 Вы были заблокированы. Причина: {reason}")
            else:
                await bot.send_message(user_id, f"✅ Вы были разблокированы. Причина: {reason}")
        except Exception as e:
            logging.error(f"Error sending block/unblock notification to user {user_id}: {e}")
        
        return RedirectResponse(url="/users", status_code=303) 

@router.get("/books/template/download")
async def download_template(request: Request):
    if not is_admin(request):
        return RedirectResponse(url="/login")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Книги"
    
    # Заголовки с пометкой обязательных полей
    headers = [
        "Название книги*", 
        "Автор*", 
        "Тематика*", 
        "Описание", 
        "Количество экземпляров*", 
        "Цена за экземпляр*", 
        "Поставщик*"
    ]
    
    # Добавляем заголовки
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Примеры данных для разных типов книг
    examples = [
        [
            "Война и мир (Том 1)", 
            "Л.Н. Толстой", 
            "Художественная литература",
            "Роман-эпопея о событиях 1812 года",
            "5",
            "1000",
            "Книжный дом"
        ],
        [
            "Алгебра 7 класс", 
            "Макарычев Ю.Н.", 
            "Учебная литература",
            "Учебник для общеобразовательных учреждений",
            "30",
            "800",
            "Просвещение"
        ],
        [
            "Гарри Поттер и философский камень", 
            "Дж. К. Роулинг", 
            "Художественная литература",
            "Первая книга о приключениях юного волшебника",
            "10",
            "700",
            "Росмэн"
        ],
        [
            "Python для начинающих", 
            "Марк Лутц", 
            "Научная литература",
            "Введение в программирование на Python",
            "15",
            "1500",
            "O'Reilly"
        ],
        [
            "Большая энциклопедия школьника", 
            "Коллектив авторов", 
            "Справочная литература",
            "Универсальный справочник для учащихся",
            "8",
            "2000",
            "АСТ"
        ]
    ]
    
    # Добавляем примеры
    for row, example in enumerate(examples, 2):
        for col, value in enumerate(example, 1):
            ws.cell(row=row, column=col, value=value)
    
    # Настраиваем ширину столбцов
    column_widths = {
        "A": 40,  # Название
        "B": 30,  # Автор
        "C": 25,  # Тематика
        "D": 50,  # Описание
        "E": 15,  # Количество
        "F": 15,  # Цена
        "G": 25   # Поставщик
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Создаем лист с инструкциями
    ws_help = wb.create_sheet(title="Инструкция")
    instructions = [
        ["Инструкция по заполнению шаблона:"],
        [""],
        ["1. Поля, отмеченные звездочкой (*), обязательны для заполнения"],
        ["2. Тематика должна быть одной из:"],
        ["   - Художественная литература"],
        ["   - Учебная литература"],
        ["   - Научная литература"],
        ["   - Справочная литература"],
        ["3. Количество экземпляров - целое положительное число"],
        ["4. Цена за экземпляр - положительное число в рублях"],
        ["5. В первом листе приведены примеры заполнения"],
        ["6. Не изменяйте структуру файла и заголовки столбцов"],
        ["7. Дата закупки будет установлена автоматически"],
        [""],
        ["Примечания:"],
        ["- Описание может быть пустым"],
        ["- Количество и цена должны быть указаны цифрами"],
        ["- Проверяйте правильность указания тематики"]
    ]
    
    for row, instruction in enumerate(instructions, 1):
        ws_help.cell(row=row, column=1, value=instruction[0])
    
    ws_help.column_dimensions["A"].width = 70
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=books_template.xlsx"}
    )

@router.post("/books/upload")
async def upload_books(request: Request, file: UploadFile = File(...)):
    if not is_admin(request):
        return RedirectResponse(url="/login")
    
    try:
        contents = await file.read()
        wb = load_workbook(BytesIO(contents))
        ws = wb.active
        
        with get_db() as conn:
            cursor = conn.cursor()
            
            # Пропускаем заголовок
            for row in ws.iter_rows(min_row=2):
                if not row[0].value:  # Пропускаем пустые строки
                    continue
                    
                title = row[0].value
                author = row[1].value
                theme = row[2].value
                description = row[3].value
                quantity = int(row[4].value or 0)
                price = float(row[5].value or 0)
                supplier = row[6].value
                
                # Добавляем книгу
                cursor.execute("""
                    INSERT INTO books (title, author, theme, description)
                    VALUES (?, ?, ?, ?)
                    RETURNING id
                """, (title, author, theme, description))
                book_id = cursor.fetchone()[0]
                
                # Добавляем закупку
                if quantity > 0 and price > 0:
                    cursor.execute("""
                        INSERT INTO book_purchases (
                            book_id, quantity, price_per_unit, 
                            supplier, purchase_date
                        )
                        VALUES (?, ?, ?, ?, datetime('now'))
                    """, (book_id, quantity, price, supplier))
                    
                    # Создаем копии книг
                    for _ in range(quantity):
                        cursor.execute("""
                            INSERT INTO book_copies (book_id, status)
                            VALUES (?, 'available')
                        """, (book_id,))
                
            conn.commit()
            
        return RedirectResponse(url="/books", status_code=303)
        
    except Exception as e:
        logging.error(f"Error uploading books: {e}")
        raise HTTPException(status_code=400, detail="Ошибка при загрузке файла") 