from database.models import get_db, log_admin_action
from datetime import datetime
import logging
from fastapi import APIRouter, Request, Form, HTTPException, Query, File, UploadFile
from fastapi.responses import RedirectResponse, JSONResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
import json
from utils.token_storage import verify_token
import qrcode
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import sqlite3
from typing import Optional
from utils.book_api import get_book_by_isbn
import os
import random
import string
import time
import aiohttp
from datetime import datetime, timedelta
from io import BytesIO
from typing import Optional, List, Dict, Any

# Создаем router с префиксом
router = APIRouter(prefix="")
templates = Jinja2Templates(directory="admin_panel/templates")
router.mount("/static", StaticFiles(directory="admin_panel/static"), name="static")

def is_admin(request: Request) -> bool:
    return request.session.get('is_admin', False)

def get_admin_info(request: Request) -> dict:
    return {
        'username': request.session.get('username', 'Admin'),
        'user_id': request.session.get('user_id')
    }

@router.get("/school")
async def school_profile_page(request: Request):
    """
    Отображает страницу с профилем школы и позволяет редактировать информацию о ней.
    """
    try:
        with get_db() as conn:
            cursor = conn.cursor()
            
            # Получаем информацию о школе
            cursor.execute("SELECT * FROM school_info LIMIT 1")
            school_data = cursor.fetchone()
            
            # Если в базе еще нет данных о школе, создаем пустой словарь
            if school_data:
                school = dict(zip([column[0] for column in cursor.description], school_data))
            else:
                school = {
                    "name": "",
                    "type": "school",
                    "address": "",
                    "phone": "",
                    "email": "",
                    "website": "",
                    "director": "",
                    "description": "",
                    "logo_url": None
                }
            
            # Получаем общую статистику
            stats = {}
            
            # Количество учащихся (роль "user")
            cursor.execute("SELECT COUNT(*) FROM users WHERE role = 'user'")
            stats["students_count"] = cursor.fetchone()[0]
            
            # Количество учителей (роль "teacher")
            cursor.execute("SELECT COUNT(*) FROM users WHERE role = 'teacher'")
            stats["teachers_count"] = cursor.fetchone()[0]
            
            # Общее количество книг
            cursor.execute("SELECT COUNT(*) FROM books")
            stats["books_count"] = cursor.fetchone()[0]
            
            # Количество выданных книг
            cursor.execute("SELECT COUNT(*) FROM borrowed_books WHERE status = 'borrowed'")
            stats["borrowed_books"] = cursor.fetchone()[0]
            
            return templates.TemplateResponse("school_profile.html", {
                "request": request,
                "school": school,
                "stats": stats
            })
    except Exception as e:
        logging.error(f"Error loading school profile page: {e}")
        return templates.TemplateResponse("error.html", {"request": request, "error": f"Ошибка загрузки профиля школы: {str(e)}"})

@router.post("/school/update-info")
async def update_school_info(request: Request):
    """
    Обновляет информацию о школе.
    """
    try:
        form_data = await request.form()
        
        # Получаем все поля из формы
        name = form_data.get("name")
        school_type = form_data.get("type")
        address = form_data.get("address")
        phone = form_data.get("phone")
        email = form_data.get("email")
        website = form_data.get("website")
        director = form_data.get("director")
        description = form_data.get("description")
        
        # Проверяем, что название школы указано
        if not name:
            return JSONResponse({"success": False, "error": "Название школы обязательно для заполнения"})
        
        with get_db() as conn:
            cursor = conn.cursor()
            
            # Проверяем, существует ли уже запись о школе
            cursor.execute("SELECT COUNT(*) FROM school_info")
            exists = cursor.fetchone()[0] > 0
            
            if exists:
                # Обновляем существующую запись
                cursor.execute("""
                    UPDATE school_info 
                    SET name = ?, type = ?, address = ?, phone = ?, 
                        email = ?, website = ?, director = ?, description = ?
                    WHERE id = 1
                """, (name, school_type, address, phone, email, website, director, description))
            else:
                # Создаем новую запись
                cursor.execute("""
                    INSERT INTO school_info 
                    (name, type, address, phone, email, website, director, description) 
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (name, school_type, address, phone, email, website, director, description))
            
            conn.commit()
            
            return JSONResponse({"success": True})
    except Exception as e:
        logging.error(f"Error updating school info: {e}")
        return JSONResponse({"success": False, "error": str(e)})

@router.post("/school/upload-logo")
async def upload_school_logo(request: Request):
    """
    Загружает логотип школы.
    """
    try:
        # Создаем директорию для хранения изображений, если её нет
        upload_dir = "static/uploads/school"
        os.makedirs(upload_dir, exist_ok=True)
        
        form = await request.form()
        logo_file = form.get("logo")
        
        if not logo_file:
            return JSONResponse({"success": False, "error": "Файл не выбран"})
        
        # Проверяем формат файла
        if not logo_file.filename.lower().endswith(('.png', '.jpg', '.jpeg', '.gif')):
            return JSONResponse({"success": False, "error": "Поддерживаются только изображения (PNG, JPG, JPEG, GIF)"})
        
        # Генерируем уникальное имя файла
        file_extension = os.path.splitext(logo_file.filename)[1]
        new_filename = f"school_logo_{uuid.uuid4().hex}{file_extension}"
        file_path = os.path.join(upload_dir, new_filename)
        
        # Сохраняем файл
        with open(file_path, "wb") as f:
            content = await logo_file.read()
            f.write(content)
            
        # Обновляем запись в базе данных с путем к логотипу
        relative_path = f"/static/uploads/school/{new_filename}"
        
        with get_db() as conn:
            cursor = conn.cursor()
            
            # Проверяем, существует ли уже запись о школе
            cursor.execute("SELECT COUNT(*) FROM school_info")
            exists = cursor.fetchone()[0] > 0
            
            if exists:
                # Получаем старый логотип, чтобы удалить его
                cursor.execute("SELECT logo_url FROM school_info WHERE id = 1")
                old_logo = cursor.fetchone()
                
                # Обновляем путь к логотипу
                cursor.execute("UPDATE school_info SET logo_url = ? WHERE id = 1", (relative_path,))
            else:
                # Создаем новую запись с логотипом
                cursor.execute("""
                    INSERT INTO school_info 
                    (name, logo_url) 
                    VALUES (?, ?)
                """, ("Моя школа", relative_path))
            
            conn.commit()
            
            # Если был старый логотип, удаляем его
            if exists and old_logo and old_logo[0]:
                try:
                    old_logo_path = old_logo[0].replace("/static", "static")
                    if os.path.exists(old_logo_path):
                        os.remove(old_logo_path)
                except Exception as e:
                    logging.error(f"Error removing old logo: {e}")
        
        return JSONResponse({"success": True})
    except Exception as e:
        logging.error(f"Error uploading school logo: {e}")
        return JSONResponse({"success": False, "error": str(e)}) 

@router.post("/books/{book_id}/edit")
async def edit_book(
    request: Request,
    book_id: int,
    title: str = Form(...),
    author: str = Form(...),
    theme: str = Form(...),
    description: str = Form(...),
    pages: str = Form(...),
    edition_number: str = Form(None),
    publication_year: str = Form(...),
    publisher: str = Form(...),
    is_textbook: str = Form('N'),
    quantity: int = Form(...)
):
    try:
        admin_id = request.session.get('user_id')
        logging.info(f"Edit book request - Session data: {dict(request.session)}")
        
        if not admin_id:
            logging.error("No admin_id in session")
            return RedirectResponse(url="/login")
        
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute("SELECT id, username, role FROM users WHERE id = ?", (admin_id,))
        user = cursor.fetchone()
        logging.info(f"User data: {user}")
        
        if not user or user[2] != 'admin':
            logging.error(f"User {admin_id} is not admin")
            return RedirectResponse(url="/login")
        
        cursor.execute("SELECT title, author FROM books WHERE id = ?", (book_id,))
        old_book = cursor.fetchone()
        logging.info(f"Book data: {old_book}")
        
        if not old_book:
            raise HTTPException(status_code=404, detail="Book not found")
        
        old_title, old_author = old_book
        
        cursor.execute("""
            UPDATE books 
            SET title = ?, author = ?, theme = ?, description = ?, pages = ?, 
                edition_number = ?, publication_year = ?, publisher = ?, quantity = ?, is_textbook = ?
            WHERE id = ?
        """, (title, author, theme, description, pages, edition_number, publication_year, publisher, quantity, is_textbook, book_id))
        
        details = f"Изменена книга ID {book_id}: '{old_title}' -> '{title}', '{old_author}' -> '{author}'"
        logging.info(f"Attempting to log action with details: {details}")
        
        log_admin_action(
            admin_id=admin_id,
            action_type="edit_book",
            book_id=book_id,
            details=details
        )
        
        conn.commit()
        
        cursor.execute("SELECT * FROM admin_logs ORDER BY id DESC LIMIT 1")
        last_log = cursor.fetchone()
        logging.info(f"Last log entry: {last_log}")
        
        return RedirectResponse(url="/books", status_code=303)
        
    except Exception as e:
        logging.error(f"Error in edit_book: {e}")
        logging.exception("Full traceback:")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals():
            conn.close()

@router.post("/books/add")
async def add_book(
    request: Request,
    title: str = Form(...),
    author: str = Form(...),
    theme: str = Form(...),
    description: str = Form(...),
    pages: str = Form(...),
    edition_number: str = Form(None),
    publication_year: str = Form(...),
    publisher: str = Form(...),
    is_textbook: str = Form('N'),
    quantity: int = Form(...)
):
    try:
        admin_id = request.session.get('user_id')
        if not admin_id: return RedirectResponse(url="/login")
        
        with get_db() as conn:
            cursor = conn.cursor()
            
            # Добавляем книгу
            cursor.execute("""
                INSERT INTO books (title, author, theme, description, pages, edition_number, publication_year, publisher, quantity, is_textbook)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (title, author, theme, description, pages, edition_number, publication_year, publisher, quantity, is_textbook))
            
            book_id = cursor.lastrowid
            conn.commit()
            
            log_admin_action(admin_id, "add_book", book_id=book_id, details=f"Добавлена книга: '{title}'")
            
            return RedirectResponse(url="/books", status_code=303)
            
    except Exception as e:
        logging.error(f"Error adding book: {e}")
        raise HTTPException(status_code=500, detail="Ошибка при добавлении книги")

@router.post("/books/{book_id}/delete")
async def delete_book(request: Request, book_id: int):
    try:
        admin_id = request.session.get('user_id')
        if not admin_id:
            return RedirectResponse(url="/login", status_code=302)
            
        with get_db() as conn:
            cursor = conn.cursor()
            
            cursor.execute("SELECT role FROM users WHERE id = ?", (admin_id,))
            result = cursor.fetchone()
            if not result or result[0] != 'admin': 
                return RedirectResponse(url="/login", status_code=302)
            
            # Получаем информацию о книге перед удалением
            cursor.execute("SELECT title, author FROM books WHERE id = ?", (book_id,))
            book = cursor.fetchone()
            if book:  
                title, author = book
                
                cursor.execute("DELETE FROM books WHERE id = ?", (book_id,))
                conn.commit()
                
                # Логируем действие
                log_admin_action(
                    admin_id=admin_id,
                    action_type="delete_book",
                    book_id=book_id,
                    details=f"Удалена книга: '{title}' by {author}"
                )
                
        return RedirectResponse(url="/books", status_code=302)
        
    except Exception as e:
        logging.error(f"Error deleting book: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/")
async def root(request: Request):
    if not is_admin(request):
        return RedirectResponse(url="/login")
    return RedirectResponse(url="/books")

@router.get("/logs")
async def logs_page(
    request: Request, 
    page: int = 1,
    admin_id: str = "",
    action_type: str = "",
    date_from: str = "",
    date_to: str = ""
):
    if not is_admin(request):
        return RedirectResponse(url="/login")
        
    admin_info = get_admin_info(request)
    items_per_page = 12
    
    try:
        with get_db() as conn:
            cursor = conn.cursor()
            
            # Получаем список админов
            cursor.execute("""
                SELECT DISTINCT u.id, u.full_name, u.username
                FROM admin_logs al
                JOIN users u ON al.admin_id = u.id
                ORDER BY COALESCE(u.full_name, u.username)
            """)
            
            admins = [{'id': row[0], 'name': row[1] or f"@{row[2]}"} for row in cursor.fetchall()]
            
            # Базовый запрос
            query = """
                SELECT 
                    al.id,
                    al.action_type,
                    al.details,
                    strftime('%d.%m.%Y %H:%M', al.timestamp) as formatted_time,
                    u.full_name,
                    u.username
                FROM admin_logs al
                JOIN users u ON al.admin_id = u.id
                WHERE 1=1
            """
            params = []
            
            # Добавляем фильтры
            if admin_id:
                query += " AND al.admin_id = ?"
                params.append(admin_id)
            
            if action_type:
                query += " AND al.action_type = ?"
                params.append(action_type)
                
            if date_from:
                try:
                    # Конвертируем дату из dd.mm.yyyy в yyyy-mm-dd
                    day, month, year = date_from.split('.')
                    formatted_date = f"{year}-{month}-{day}"
                    query += " AND date(al.timestamp) >= date(?)"
                    params.append(formatted_date)
                except:
                    pass
                
            if date_to:
                try:
                    # Конвертируем дату из dd.mm.yyyy в yyyy-mm-dd
                    day, month, year = date_to.split('.')
                    formatted_date = f"{year}-{month}-{day}"
                    query += " AND date(al.timestamp) <= date(?)"
                    params.append(formatted_date)
                except:
                    pass
            
            # Получаем общее количество записей
            count_query = f"SELECT COUNT(*) FROM ({query})"
            cursor.execute(count_query, params)
            total_records = cursor.fetchone()[0]
            
            # Вычисляем пагинацию
            total_pages = (total_records + items_per_page - 1) // items_per_page
            
            # Проверяем, что page находится в допустимом диапазоне
            page = max(1, min(page, total_pages)) if total_pages > 0 else 1
            offset = (page - 1) * items_per_page
            
            # Добавляем сортировку и лимиты
            query += " ORDER BY al.timestamp DESC LIMIT ? OFFSET ?"
            params.extend([items_per_page, offset])
            
            # Получаем записи
            logging.info(f"SQL Query: {query}")
            logging.info(f"Params: {params}")
            cursor.execute(query, params)
            
            logs = []
            for row in cursor.fetchall():
                logs.append({
                    'id': row[0],
                    'action_type': row[1],
                    'details': row[2],
                    'timestamp': row[3],
                    'admin_name': row[4] or f"@{row[5]}"
                })
            
            # Формируем строку фильтров для пагинации
            filter_parts = []
            if admin_id:
                filter_parts.append(f"admin_id={admin_id}")
            if action_type:
                filter_parts.append(f"action_type={action_type}")
            if date_from:
                filter_parts.append(f"date_from={date_from}")
            if date_to:
                filter_parts.append(f"date_to={date_to}")
            filter_query = "&".join(filter_parts)
                
            return templates.TemplateResponse(
                "logs.html",
                {
                    "request": request,
                    "logs": logs,
                    "admin_info": admin_info,
                    "current_page": page,
                    "total_pages": total_pages,
                    "filter_query": filter_query,
                    "filters": {
                        "admin_id": admin_id,
                        "action_type": action_type,
                        "date_from": date_from,
                        "date_to": date_to
                    },
                    "admins": admins
                }
            )
    except Exception as e:
        logging.error(f"Error loading logs: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/login")
async def login_page(request: Request):
    token = request.query_params.get('token')
    
    if token:
        # Проверяем токен
        user_id = verify_token(token)
        if user_id:
            # Если токен валидный, сохраняем данные в сессию
            request.session['user_id'] = user_id
            request.session['is_admin'] = True
            
            # Перенаправляем на главную страницу
            return RedirectResponse(url="/books")
            
    # Если нет токена или токен невалидный, показываем ошибку
    return templates.TemplateResponse(
        "error.html",
        {
            "request": request,
            "message": "Недействительная или истекшая ссылка. Пожалуйста, получите новую ссылку через бота."
        }
    )

@router.get("/books")
async def books_page(request: Request, page: int = Query(1, ge=1), search: str = Query(""), sort: str = Query("title"), order: str = Query("asc")):
    try:
        with get_db() as conn:
            cursor = conn.cursor()
            
            # Параметры пагинации
            per_page = 15
            offset = (page - 1) * per_page
            
            # Проверка и валидация параметров сортировки
            valid_sort_fields = {
                'id': 'b.id', 
                'title': 'b.title',
                'author': 'b.author',
                'theme': 'b.theme',
                'pages': 'b.pages',
                'edition_number': 'b.edition_number',
                'publication_year': 'b.publication_year',
                'publisher': 'b.publisher',
                'available_copies': "(b.quantity - COUNT(DISTINCT CASE WHEN bb.status = 'borrowed' THEN bb.copy_id END))",
                'total_copies': 'b.quantity'
            }
            
            sort_field = valid_sort_fields.get(sort, 'b.title')  # По умолчанию сортировка по названию
            sort_order = "ASC" if order.lower() == "asc" else "DESC"
            
            # Базовый запрос для получения книг
            base_query = """
                SELECT 
                    b.id, b.title, b.author, b.theme, b.description, b.pages, 
                    b.edition_number, b.publication_year, b.publisher,
                    b.quantity as total_copies,
                    (b.quantity - COUNT(DISTINCT CASE WHEN bb.status = 'borrowed' THEN bb.copy_id END)) as available_copies,
                    0 as avg_price, -- Убрали расчет средней цены из-за отсутствия таблицы purchases
                    b.is_textbook -- Добавляем поле is_textbook
                FROM books b
                LEFT JOIN book_copies bc ON b.id = bc.book_id
                LEFT JOIN borrowed_books bb ON bc.id = bb.copy_id
                -- LEFT JOIN purchases p ON b.id = p.book_id -- Убрали соединение
            """
            
            # Условия поиска
            conditions = []
            params = []
            
            if search:
                conditions.append("(b.title LIKE ? OR b.author LIKE ?)")
                search_param = f"%{search}%"
                params.extend([search_param, search_param])
            
            # Добавляем условия к запросу
            if conditions:
                base_query += " WHERE " + " AND ".join(conditions)
            
            # Группировка и сортировка
            base_query += """
                GROUP BY b.id, b.title, b.author, b.theme, b.description, b.pages, 
                         b.edition_number, b.publication_year, b.publisher, b.quantity, b.is_textbook -- Добавили is_textbook в GROUP BY
                ORDER BY {sort_field} {sort_order}
                LIMIT ? OFFSET ?
            """.format(sort_field=sort_field, sort_order=sort_order)
            
            # Добавляем параметры пагинации
            params.extend([per_page, offset])
            
            # Выполняем запрос на получение книг
            cursor.execute(base_query, params)
            books_data = cursor.fetchall()
            
            # Преобразуем результат в список словарей
            books = [dict(zip([column[0] for column in cursor.description], row)) for row in books_data]
            
            # Запрос для получения общего количества книг (с учетом фильтра)
            count_query = "SELECT COUNT(DISTINCT b.id) FROM books b"
            count_params = []
            if search:
                count_query += " WHERE (b.title LIKE ? OR b.author LIKE ?)"
                count_params.extend([search_param, search_param])
                
            cursor.execute(count_query, count_params)
            total_books_filtered = cursor.fetchone()[0]
            
            # Вычисляем общее количество страниц
            total_pages = (total_books_filtered + per_page - 1) // per_page if total_books_filtered > 0 else 1
            
            # Получаем общую статистику (без фильтра)
            cursor.execute("SELECT COUNT(*) FROM books")
            total_books_all = cursor.fetchone()[0]
            cursor.execute("SELECT SUM(quantity) FROM books")
            total_copies_all = cursor.fetchone()[0] or 0
            cursor.execute("SELECT COUNT(DISTINCT copy_id) FROM borrowed_books WHERE status = 'borrowed'")
            borrowed_copies_all = cursor.fetchone()[0] or 0
            
            # Получаем количество просроченных книг
            cursor.execute("""
                SELECT COUNT(DISTINCT bb.copy_id) 
                FROM borrowed_books bb 
                WHERE bb.status = 'borrowed' AND bb.return_date < date('now')
            """)
            overdue_copies_all = cursor.fetchone()[0] or 0
            
            return templates.TemplateResponse("books.html", {
                "request": request,
                "books": books,
                "total_books": total_books_all,
                "total_copies": total_copies_all,
                "borrowed_copies": borrowed_copies_all,
                "overdue_copies": overdue_copies_all,
                "current_page": page,
                "total_pages": total_pages,
                "search_query": search,
                "sort": sort,
                "order": order,
                "max": max,
                "min": min
            })
            
    except Exception as e:
        logging.error(f"Error loading books page: {e}")
        return templates.TemplateResponse("error.html", {"request": request, "error": "Ошибка загрузки страницы книг"})

@router.get("/suggestions")
async def suggestions_page(request: Request):
    if not is_admin(request):
        return RedirectResponse(url="/login")
    
    admin_info = get_admin_info(request)
    
    try:
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT 
                    bs.id,
                    bs.title,
                    bs.url,
                    bs.price,
                    bs.reason,
                    bs.status,
                    bs.created_at,
                    u.full_name,
                    u.username
                FROM book_suggestions bs
                JOIN users u ON bs.user_id = u.id
                ORDER BY bs.created_at DESC
            """)
            
            suggestions = []
            for row in cursor.fetchall():
                suggestions.append({
                    'id': row[0],
                    'title': row[1],
                    'url': row[2],
                    'price': row[3],
                    'reason': row[4],
                    'status': row[5],
                    'created_at': row[6],
                    'user_name': row[7] or f"@{row[8]}"
                })
                
            return templates.TemplateResponse(
                "suggestions.html",
                {
                    "request": request, 
                    "suggestions": suggestions,
                    "admin_info": admin_info
                }
            )
    except Exception as e:
        logging.error(f"Error loading suggestions: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/suggestions/{suggestion_id}/status")
async def update_suggestion_status(suggestion_id: int, request: Request):
    if not is_admin(request):
        return JSONResponse({"success": False, "error": "Unauthorized"})
        
    try:
        data = await request.json()
        new_status = data.get('status')
        
        if new_status not in ['approved', 'rejected']:
            return JSONResponse({"success": False, "error": "Invalid status"})
            
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE book_suggestions 
                SET status = ? 
                WHERE id = ?
            """, (new_status, suggestion_id))
            conn.commit()
            
        return JSONResponse({"success": True})
        
    except Exception as e:
        logging.error(f"Error updating suggestion status: {e}")
        return JSONResponse({"success": False, "error": str(e)})

@router.get("/books/{book_id}/qrcodes")
async def book_qrcodes(request: Request, book_id: int):
    if not is_admin(request):
        return RedirectResponse(url="/login")
        
    admin_info = get_admin_info(request)
    
    try:
        with get_db() as conn:
            cursor = conn.cursor()
            
            # Получаем информацию о книге
            cursor.execute("SELECT title, author FROM books WHERE id = ?", (book_id,))
            book = cursor.fetchone()
            if not book:
                raise HTTPException(status_code=404, detail="Book not found")
                
            # Получаем QR коды книги с правильным статусом
            cursor.execute("""
                SELECT 
                    bc.id, 
                    CASE 
                        WHEN bb.status = 'borrowed' THEN 'borrowed'
                        WHEN bc.status = 'written_off' THEN 'written_off'
                        ELSE 'available'
                    END as status,
                    bb.user_id,
                    u.full_name,
                    u.username
                FROM book_copies bc
                LEFT JOIN borrowed_books bb ON bc.id = bb.copy_id 
                    AND bb.status = 'borrowed'
                LEFT JOIN users u ON bb.user_id = u.id
                WHERE bc.book_id = ?
                ORDER BY bc.id
            """, (book_id,))
            
            copies = []
            for row in cursor.fetchall():
                copy_id, status, user_id, full_name, username = row
                copies.append({
                    'id': copy_id,
                    'status': status,
                    'borrowed_by': full_name or f"@{username}" if user_id else None
                })
                
            return templates.TemplateResponse(
                "book_qrcodes.html",
                {
                    "request": request,
                    "book": {
                        'id': book_id,
                        'title': book[0],
                        'author': book[1]
                    },
                    "copies": copies,
                    "admin_info": admin_info
                }
            )
    except Exception as e:
        logging.error(f"Error loading QR codes: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/books/{book_id}/purchases")
async def get_book_purchases(request: Request, book_id: int):
    if not is_admin(request):
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    try:
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT 
                    strftime('%d.%m.%Y', purchase_date) as formatted_date,
                    quantity,
                    price_per_unit,
                    supplier
                FROM book_purchases
                WHERE book_id = ?
                ORDER BY purchase_date DESC
            """, (book_id,))
            
            purchases = []
            for row in cursor.fetchall():
                purchases.append({
                    'date': row[0],
                    'quantity': row[1],
                    'price': row[2],
                    'supplier': row[3]
                })
                
            return JSONResponse({"purchases": purchases})
    except Exception as e:
        logging.error(f"Error getting purchases: {e}")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.post("/books/{book_id}/purchases/add")
async def add_book_purchase(request: Request, book_id: int):
    if not is_admin(request):
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
        
    try:
        data = await request.json()
        quantity = data.get('quantity')
        price = data.get('price')
        supplier = data.get('supplier')
        admin_id = request.session.get('user_id')
        
        with get_db() as conn:
            cursor = conn.cursor()
            
            # Начинаем транзакцию
            cursor.execute("BEGIN")
            
            try:
                # Получаем информацию о книге
                cursor.execute("SELECT title FROM books WHERE id = ?", (book_id,))
                book = cursor.fetchone()
                if not book:
                    return JSONResponse({"error": "Книга не найдена"}, status_code=404)
                
                book_title = book[0]
                
                # Добавляем запись о закупке с текущей датой
                cursor.execute("""
                    INSERT INTO book_purchases (book_id, quantity, price_per_unit, supplier, purchase_date)
                    VALUES (?, ?, ?, ?, datetime('now', 'localtime'))
                """, (book_id, quantity, price, supplier))
                
                # Получаем текущее количество копий для этой книги
                cursor.execute("SELECT COUNT(*) FROM book_copies WHERE book_id = ?", (book_id,))
                current_copies = cursor.fetchone()[0]
                
                # Создаем новые копии книги
                for i in range(quantity):
                    copy_number = current_copies + i + 1
                    cursor.execute("""
                        INSERT INTO book_copies (book_id, copy_number, status)
                        VALUES (?, ?, 'available')
                    """, (book_id, copy_number))
                
                # Обновляем количество книг
                cursor.execute("""
                    UPDATE books 
                    SET quantity = quantity + ?
                    WHERE id = ?
                """, (quantity, book_id))
                
                # Логируем действие в той же транзакции
                cursor.execute("""
                    INSERT INTO admin_logs (admin_id, action_type, book_id, details, timestamp)
                    VALUES (?, ?, ?, ?, datetime('now', 'localtime'))
                """, (
                    admin_id,
                    "add_purchase",
                    book_id,
                    f"Добавлена закупка книги '{book_title}': {quantity} шт. по {price} ₽ от {supplier}"
                ))
                
                # Завершаем транзакцию
                conn.commit()
                return JSONResponse({"success": True})
                
            except Exception as e:
                # В случае ошибки откатываем изменения
                cursor.execute("ROLLBACK")
                raise e
            
    except Exception as e:
        logging.error(f"Error adding purchase: {e}")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.get("/qr/{copy_id}")
async def get_qr_code(copy_id: int):
    # Генерируем QR код
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(str(copy_id))
    qr.make(fit=True)
    
    img = qr.make_image(fill_color="black", back_color="white")
    
    # Сохраняем в буфер
    buf = BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)
    
    return StreamingResponse(buf, media_type="image/png")

@router.post("/books/copies/{copy_id}/write-off")
async def write_off_copy(request: Request, copy_id: int):
    if not is_admin(request):
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
        
    try:
        with get_db() as conn:
            cursor = conn.cursor()
            
            # Начинаем транзакцию
            cursor.execute("BEGIN")
            
            try:
                # Проверяем существование и статус копии
                cursor.execute("""
                    SELECT bc.id, b.title, b.id 
                    FROM book_copies bc
                    JOIN books b ON bc.book_id = b.id
                    WHERE bc.id = ? AND bc.status = 'available'
                """, (copy_id,))
                
                result = cursor.fetchone()
                if not result:
                    return JSONResponse(
                        {"error": "Копия не найдена или уже выдана"}, 
                        status_code=404
                    )
                
                book_id = result[2]
                
                # Списываем копию
                cursor.execute("""
                    UPDATE book_copies 
                    SET status = 'written_off' 
                    WHERE id = ?
                """, (copy_id,))
                
                # Уменьшаем количество книг
                cursor.execute("""
                    UPDATE books 
                    SET quantity = quantity - 1
                    WHERE id = ?
                """, (book_id,))
                
                # Добавляем запись в лог
                admin_id = request.session.get('user_id')
                cursor.execute("""
                    INSERT INTO admin_logs (admin_id, action_type, book_id, details, timestamp)
                    VALUES (?, ?, ?, ?, datetime('now', 'localtime'))
                """, (admin_id, "write_off_copy", book_id, f"Списан экземпляр #{copy_id} книги '{result[1]}'"))
                
                # Завершаем транзакцию
                conn.commit()
                return JSONResponse({"success": True})
                
            except Exception as e:
                # Если произошла ошибка, откатываем изменения
                cursor.execute("ROLLBACK")
                raise e
            
    except Exception as e:
        logging.error(f"Error writing off copy: {e}")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.post("/api/books")
async def create_book(request: Request):
    if not is_admin(request):
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    try:
        data = await request.json()
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO books (title, author, description, theme, quantity)
                VALUES (?, ?, ?, ?, 0)
            """, (data['title'], data['author'], data['description'], data['theme']))
            conn.commit()
        return JSONResponse({"success": True})
    except Exception as e:
        logging.error(f"Error creating book: {e}")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.get("/users")
async def users_page(request: Request, 
                    page: int = Query(1, ge=1),
                    search: str = Query(""),
                    role: str = Query(""),
                    status: str = Query(""),
                    sort: str = Query("id"),
                    order: str = Query("desc")):
    logging.info(f"Received filter params: page={page}, search={search}, role={role}, status={status}, sort={sort}, order={order}")
    try:
        with get_db() as conn:
            cursor = conn.cursor()
            
            # Проверка и валидация параметров сортировки
            valid_sort_fields = {
                'id': 'u.id',
                'username': 'u.username',
                'full_name': 'u.full_name',
                'phone': 'u.phone',
                'role': 'u.role',
                'class': 'u.class',
                'is_blocked': 'u.is_blocked',
                'borrow_count': '0', # Заглушка, так как это вычисляемое поле
                'review_count': '0'  # Заглушка, так как это вычисляемое поле
            }
            
            sort_field = valid_sort_fields.get(sort, 'u.id')  # По умолчанию сортировка по id
            sort_order = "ASC" if order.lower() == "asc" else "DESC"
            
            # Базовый запрос - максимально упрощенный
            query = """
                SELECT u.id, u.username, u.full_name, u.phone, u.role,
                      0 as borrow_count,
                      0 as review_count,
                      '' as last_activity,
                      u.is_blocked,
                      u.class
                FROM users u
            """
            
            # Добавление условий фильтрации
            conditions = []
            params = []
            
            if search:
                conditions.append("(u.username LIKE ? OR u.full_name LIKE ? OR u.id LIKE ?)")
                search_param = f"%{search}%"
                params.extend([search_param, search_param, search_param])
            
            if role:
                conditions.append("u.role = ?")
                params.append(role)
            
            if status == "active":
                conditions.append("u.is_blocked = 0")
            elif status == "blocked":
                conditions.append("u.is_blocked = 1")
            
            if conditions:
                query += " WHERE " + " AND ".join(conditions)
            
            # Добавляем сортировку и пагинацию
            query += f"""
                ORDER BY {sort_field} {sort_order}
                LIMIT ? OFFSET ?
            """
            
            # Добавляем параметры для пагинации
            per_page = 20
            offset = (page - 1) * per_page
            params.extend([per_page, offset])
            
            # Выполняем запрос
            logging.info(f"SQL Query: {query}")
            logging.info(f"Params: {params}")
            cursor.execute(query, params)
            users = cursor.fetchall()
            
            # Получаем общее количество пользователей для пагинации
            count_query = """
                SELECT COUNT(DISTINCT u.id)
                FROM users u
            """
            
            if conditions:
                count_query += " WHERE " + " AND ".join(conditions)
            
            cursor.execute(count_query, params[:-2])  # Исключаем параметры пагинации
            total_users = cursor.fetchone()[0]
            
            # Получаем статистику
            cursor.execute("""
                SELECT COUNT(*) as total,
                       SUM(CASE WHEN role = 'admin' THEN 1 ELSE 0 END) as admins,
                       SUM(CASE WHEN role = 'teacher' THEN 1 ELSE 0 END) as teachers,
                       SUM(CASE WHEN role = 'user' THEN 1 ELSE 0 END) as students
                FROM users
            """)
            stats = cursor.fetchone()
            
            # Вычисляем количество страниц
            if total_users > 0 and per_page > 0:
                total_pages = (total_users + per_page - 1) // per_page
            else:
                total_pages = 1 # Устанавливаем 1, если пользователей нет или ошибка
            
            # Явно передаем встроенные функции в шаблон
            template_context = {
                "request": request,
                "users": users,
                "stats": stats,
                "current_page": page,
                "total_pages": total_pages,
                "filters": {
                    "search": search,
                    "role": role,
                    "status": status
                },
                "sort": sort,
                "order": order,
                "max": max, # Передаем функцию max
                "min": min  # Передаем функцию min
            }
            
            return templates.TemplateResponse("users.html", template_context)
            
    except Exception as e:
        logging.error(f"Error in users_page: {e}")
        return templates.TemplateResponse("error.html", {
            "request": request,
            "error": "Произошла ошибка при загрузке страницы пользователей"
        })

@router.post("/users/{user_id}/toggle_block")
async def toggle_user_block(request: Request, user_id: int, reason: str = Form(...)):
    if not is_admin(request):
        return RedirectResponse(url="/login")
        
    try:
        with get_db() as conn:
            cursor = conn.cursor()
            
            # Получаем текущий статус блокировки
            cursor.execute("SELECT is_blocked FROM users WHERE id = ?", (user_id,))
            result = cursor.fetchone()
            if not result:
                raise HTTPException(status_code=404, detail="Пользователь не найден")
                
            current_status = result[0]
            new_status = 0 if current_status == 1 else 1
            
            # Обновляем статус
            cursor.execute("""
                UPDATE users 
                SET is_blocked = ?
                WHERE id = ?
            """, (new_status, user_id))
            
            conn.commit()
            
            # Логируем действие
            admin_id = request.session.get('user_id')
            action = "unblock_user" if new_status == 0 else "block_user"
            log_admin_action(
                admin_id=admin_id,
                action_type=action,
                details=f"{'Разблокирован' if new_status == 0 else 'Заблокирован'} пользователь {user_id}. Причина: {reason}"
            )
            
        return RedirectResponse(url="/users", status_code=302)
        
    except Exception as e:
        logging.error(f"Error toggling user block: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/reviews")
async def reviews_page(request: Request, search: str = "", rating: str = "", status: str = ""):
    if not is_admin(request):
        return RedirectResponse(url="/login")
        
    admin_info = get_admin_info(request)
    
    try:
        with get_db() as conn:
            cursor = conn.cursor()
            
            # Получаем статистику отзывов
            cursor.execute("""
                SELECT 
                    COUNT(*) as total_reviews,
                    SUM(CASE WHEN rating >= 4 THEN 1 ELSE 0 END) as positive_reviews,
                    SUM(CASE WHEN rating <= 2 THEN 1 ELSE 0 END) as negative_reviews
                FROM book_reviews
            """)
            stats = cursor.fetchone()
            
            # Базовый запрос для отзывов
            query = """
                SELECT 
                    br.id,
                    br.rating,
                    br.review_text,
                    br.status,
                    strftime('%d.%m.%Y %H:%M', br.created_at) as created_at,
                    u.username,
                    u.full_name,
                    b.title as book_title
                FROM book_reviews br
                JOIN users u ON br.user_id = u.id
                JOIN books b ON br.book_id = b.id
                WHERE 1=1
            """
            params = []
            
            # Добавляем условия поиска
            if search:
                query += """ 
                    AND (
                        u.username LIKE ? OR 
                        u.full_name LIKE ? OR 
                        b.title LIKE ? OR
                        br.review_text LIKE ?
                    )
                """
                search_param = f"%{search}%"
                params.extend([search_param, search_param, search_param, search_param])
            
            # Фильтр по рейтингу
            if rating:
                if rating == 'positive':
                    query += " AND br.rating >= 4"
                elif rating == 'negative':
                    query += " AND br.rating <= 2"
                elif rating == 'neutral':
                    query += " AND br.rating = 3"
            
            # Фильтр по статусу
            if status:
                query += " AND br.status = ?"
                params.append(status)
            
            # Добавляем сортировку
            query += " ORDER BY br.created_at DESC"
            
            cursor.execute(query, params)
            rows = cursor.fetchall()  # Получаем результаты один раз
            
            # Преобразуем результаты в словари
            formatted_reviews = []
            for row in rows:  # Используем полученные ранее результаты
                formatted_reviews.append({
                    'id': row[0],
                    'rating': row[1],
                    'review_text': row[2],
                    'status': row[3],
                    'created_at': row[4],
                    'username': row[5],
                    'full_name': row[6],
                    'book_title': row[7]
                })
            
            return templates.TemplateResponse(
                "reviews.html",
                {
                    "request": request,
                    "admin_info": admin_info,
                    "stats": stats,
                    "reviews": formatted_reviews,
                    "filters": {
                        "search": search,
                        "rating": rating,
                        "status": status
                    }
                }
            )
            
    except Exception as e:
        logging.error(f"Error loading reviews: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/reviews/{review_id}/{action}")
async def handle_review(request: Request, review_id: int, action: str):
    if not is_admin(request):
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
        
    if action not in ['approve', 'reject']:
        return JSONResponse({"error": "Invalid action"}, status_code=400)
        
    try:
        with get_db() as conn:
            cursor = conn.cursor()
            
            # Обновляем статус отзыва
            cursor.execute("""
                UPDATE book_reviews 
                SET status = ?
                WHERE id = ?
            """, (action + 'd', review_id))  # approved или rejected
            
            conn.commit()
            
            return JSONResponse({"success": True})
            
    except Exception as e:
        logging.error(f"Error handling review: {e}")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.get("/books/{book_id}/qrcodes")
async def book_qrcodes_page(request: Request, book_id: int):
    if not is_admin(request):
        return RedirectResponse(url="/login")
        
    admin_info = get_admin_info(request)
    
    try:
        with get_db() as conn:
            cursor = conn.cursor()
            
            # Получаем информацию о книге
            cursor.execute("""
                SELECT id, title, author 
                FROM books
                WHERE id = ?
            """, (book_id,))
            book = cursor.fetchone()
            
            if not book:
                raise HTTPException(status_code=404, detail="Книга не найдена")
                
            book_info = {
                "id": book[0],
                "title": book[1],
                "author": book[2]
            }
            
            # Получаем все экземпляры книги
            cursor.execute("""
                SELECT id, copy_number, status
                FROM book_copies
                WHERE book_id = ?
                ORDER BY copy_number
            """, (book_id,))
            
            copies = []
            for row in cursor.fetchall():
                copies.append({
                    "id": row[0],
                    "copy_number": row[1],
                    "status": row[2]
                })
            
            return templates.TemplateResponse(
                "book_qrcodes.html",
                {
                    "request": request,
                    "admin_info": admin_info,
                    "book": book_info,
                    "copies": copies
                }
            )
    except Exception as e:
        logging.error(f"Error loading book QR codes page: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/admin/books/template/download")
async def download_excel_template(request: Request):
    if not is_admin(request):
        return RedirectResponse(url="/login")
    
    # Создаем Excel файл
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Книги"
    
    # Определяем заголовки для полного шаблона (без ISBN*)
    headers = [
        "ISBN", "Название книги*", "Автор*", "Тематика(Жанр)*", "Описание*",
        "Кол-во страниц*", "Номер издания (если есть)", "Год издания*", "Издательство*",
        "Дата поставки*", "Количество экземпляров*", "Цена за экземпляр*", "Поставщик*",
        "Учебник(Y или N)"
    ]
    
    # Заполняем заголовки
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = ws[f"{col_letter}1"]
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[col_letter].width = 20
    
    # Пояснения (добавляем строку с пояснениями)
    explanation_row = [
        "Необязательно",
        "ОБЯЗАТЕЛЬНО", 
        "ОБЯЗАТЕЛЬНО", 
        "ОБЯЗАТЕЛЬНО", 
        "ОБЯЗАТЕЛЬНО",
        "ОБЯЗАТЕЛЬНО (число)", 
        "Необязательно", 
        "ОБЯЗАТЕЛЬНО (год)",
        "ОБЯЗАТЕЛЬНО",
        "ОБЯЗАТЕЛЬНО (дд.мм.гггг)",
        "ОБЯЗАТЕЛЬНО (число)",
        "ОБЯЗАТЕЛЬНО (число)",
        "ОБЯЗАТЕЛЬНО",
        "Y или N"
    ]
    
    for col_num, value in enumerate(explanation_row, 1):
        col_letter = get_column_letter(col_num)
        cell = ws[f"{col_letter}2"]
        cell.value = value
        cell.font = Font(italic=True, color="FF0000")
    
    # Добавляем пример данных (рабочий пример)
    example_row = [
        "978-5-17-087063-9",    # ISBN (необязательно)
        "Война и мир",          # Название книги (обязательно)
        "Лев Толстой",          # Автор (обязательно)
        "Классика",             # Тематика (обязательно)
        "Роман-эпопея о событиях 1812 года", # Описание (обязательно)
        "1300",                  # Страницы (обязательно)
        "1",                     # Номер издания (необязательно)
        "1869",                  # Год издания (обязательно)
        "Русский вестник",       # Издательство (обязательно)
        datetime.now().strftime("%d.%m.%Y"), # Дата поставки (обязательно)
        "30",                    # Количество (обязательно)
        "1000",                  # Цена (обязательно)
        "Книжный дом",           # Поставщик (обязательно)
        "N"                      # Учебник (Y/N)
    ]
    
    for col_num, value in enumerate(example_row, 1):
        col_letter = get_column_letter(col_num)
        cell = ws[f"{col_letter}3"]
        cell.value = value
    
    # Комментарии к ячейкам
    ws['A1'].comment = openpyxl.comments.Comment('ISBN необязателен, если заполнены все поля книги', 'Template')
    ws['B1'].comment = openpyxl.comments.Comment('Название книги обязательно', 'Template')
    ws['J1'].comment = openpyxl.comments.Comment('Формат даты: ДД.ММ.ГГГГ', 'Template')
    ws['K1'].comment = openpyxl.comments.Comment('Только целое число', 'Template')
    ws['N1'].comment = openpyxl.comments.Comment('Y - учебник, N - не учебник', 'Template')
    
    # Сохраняем в буфер
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Возвращаем файл
    response_headers = {
        "Content-Disposition": "attachment; filename=books_template.xlsx"
    }
    return StreamingResponse(
        buffer, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=response_headers
    )

@router.get("/admin/books/template/isbn/download")
async def download_excel_isbn_template(request: Request):
    if not is_admin(request):
        return RedirectResponse(url="/login")
    
    # Создаем Excel файл
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Закупки по ISBN"
    
    # Определяем заголовки
    headers = [
        "ISBN*", "Дата поставки*", "Количество экземпляров*", "Цена за экземпляр*", "Поставщик*"
    ]
    
    # Форматирование и заполнение шаблона
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = ws[f"{col_letter}1"]
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[col_letter].width = 20
    
    # Добавляем пример данных
    example_data = [
        "9780141036144", "01.01.2023", "10", "500", "ООО Издательство"
    ]
    
    for col_num, value in enumerate(example_data, 1):
        ws[f"{get_column_letter(col_num)}2"] = value
    
    # Добавляем пояснения
    ws.insert_rows(2)
    ws["A2"] = "ОБЯЗАТЕЛЬНО"
    ws["B2"] = "В формате ДД.ММ.ГГГГ"
    ws["C2"] = "Целое число"
    ws["D2"] = "Только цифры"
    ws["E2"] = "Текст"
    
    # Добавляем информацию о новом функционале получения данных по ISBN
    ws.append([])
    ws.append(["Информация:", "Данные о книге будут автоматически получены через API Google Books по указанному ISBN"])
    ws.append(["", "Если книга с таким ISBN уже существует, будет добавлена только закупка"])
    ws.append(["", "Если книга не найдена в базе, будет создана новая на основе данных из API"])
    
    # Объединяем ячейки для пояснений
    for row in range(5, 8):
        ws.merge_cells(f"B{row}:E{row}")
        ws[f"B{row}"].alignment = Alignment(horizontal="left", vertical="center")
    
    # Выделяем информационные строки
    for row in range(5, 8):
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        ws[f"B{row}"].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    
    # Сохраняем в буфер
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Возвращаем файл
    return StreamingResponse(
        buffer, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=books_isbn_template.xlsx"}
    )

# Функция для получения данных о книге по ISBN через Google Books API
async def fetch_book_by_isbn(isbn: str) -> Optional[Dict[str, Any]]:
    """
    Получает данные о книге по ISBN через Google Books API
    
    Args:
        isbn: ISBN книги для поиска
        
    Returns:
        словарь с данными о книге или None, если книга не найдена
    """
    logging.info(f"Получение данных о книге по ISBN: {isbn}")
    url = f"https://www.googleapis.com/books/v1/volumes?q=isbn:{isbn}"
    
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(url) as response:
                if response.status != 200:
                    logging.error(f"Ошибка API: {response.status} - {await response.text()}")
                    return None
                
                data = await response.json()
                if data.get('totalItems', 0) == 0:
                    logging.warning(f"Книга с ISBN {isbn} не найдена в Google Books API")
                    return None
                
                # Получаем первый результат (наиболее релевантный)
                volume_info = data['items'][0]['volumeInfo']
                
                # Извлекаем нужные данные
                book_data = {
                    'title': volume_info.get('title', ''),
                    'author': ', '.join(volume_info.get('authors', ['Неизвестный автор'])),
                    'description': volume_info.get('description', 'Описание отсутствует'),
                    'publisher': volume_info.get('publisher', 'Неизвестно'),
                    'pages': volume_info.get('pageCount', 0),
                    'publication_year': volume_info.get('publishedDate', '')[:4] if volume_info.get('publishedDate') else '',
                    'isbn': isbn
                }
                
                # Определяем тематику на основе категорий если они есть
                categories = volume_info.get('categories', [])
                book_data['theme'] = categories[0] if categories else "Разное"
                
                logging.info(f"Успешно получены данные о книге: {book_data['title']} от {book_data['author']}")
                return book_data
                
    except Exception as e:
        logging.error(f"Ошибка при получении данных о книге по ISBN {isbn}: {e}")
        return None

@router.post("/admin/books/upload")
async def upload_excel(request: Request, file: UploadFile = File(...)):
    if not is_admin(request):
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    try:
        admin_info = get_admin_info(request)
        admin_id = admin_info.get('user_id', 1)  # Используем ID администратора для логирования
        
        # Получаем режим загрузки (full или isbn)
        form = await request.form()
        mode = form.get('mode', 'full')
        file = form.get('file')
        
        if not file:
            return JSONResponse({"error": "Файл не найден"}, status_code=400)
        
        contents = await file.read()
        wb = openpyxl.load_workbook(BytesIO(contents), data_only=True)
        ws = wb.active

        # Проверяем заголовки
        headers = [str(cell.value or "").strip() for cell in ws[1]]  # Получаем заголовки из первой строки
        
        # В зависимости от режима определяем требуемые заголовки
        required_purchase_headers = ["Дата поставки*", "Количество экземпляров*", "Цена за экземпляр*", "Поставщик*"]
        
        # Явно устанавливаем режим работы в зависимости от параметра mode
        is_isbn_required = False
        is_title_required = False
        
        if mode == 'isbn':
            # Режим загрузки закупок по ISBN с получением данных через API
            is_isbn_required = True
            required_headers = ["ISBN*"] + required_purchase_headers
            logging.info(f"Явный режим ISBN с API: Обязательные поля: {required_headers}")
        else:
            # Режим загрузки полных данных книг
            is_title_required = True
            required_book_headers = ["Название книги*", "Автор*", "Тематика(Жанр)*", "Описание*",
                                   "Кол-во страниц*", "Год издания*", "Издательство*"]
            required_headers = required_book_headers + required_purchase_headers
            logging.info(f"Явный режим полных данных: Обязательные поля: {required_headers}")
        
        # Проверяем наличие всех нужных заголовков
        missing_headers = [header for header in required_headers if header not in headers]
        if missing_headers:
            return JSONResponse({
                "error": f"Отсутствуют обязательные столбцы: {', '.join(missing_headers)}. Пожалуйста, используйте шаблон."
            }, status_code=400)

        # Находим индексы столбцов
        header_indices = {}
        for i, header in enumerate(headers):
            if header:  # Пропускаем пустые заголовки
                header_indices[header] = i

        with get_db() as conn:
            cursor = conn.cursor()
            processed_books = 0
            added_purchases = 0
            skipped_rows = 0
            error_rows = []
            row_errors_details = []

            logging.info("Начало обработки Excel файла для загрузки книг и закупок.")

            # Пропускаем строку с пояснениями, если она есть
            start_row = 2
            if ws.max_row > 2:
                second_row_values = [cell.value for cell in ws[2]]
                if any(str(val).upper().startswith("ОБЯЗАТЕЛЬНО") for val in second_row_values if val):
                    start_row = 3
                    logging.info("Обнаружена строка с пояснениями, начинаем со строки 3")

            for i, row_cells in enumerate(ws.iter_rows(min_row=start_row, max_row=ws.max_row), start_row):
                row_values = [cell.value for cell in row_cells]
                logging.info(f"--- Обработка строки {i} ---")

                # Пропускаем полностью пустые строки
                if all(value is None or value == "" for value in row_values):
                    logging.info(f"Строка {i}: Пропущена (пустая).")
                    continue

                book_id = None
                row_error_messages = []
                try:
                    # Извлекаем данные в зависимости от режима работы (ISBN или полные данные книги)
                    isbn = None
                    title = None
                    author = None
                    theme = None
                    description = None
                    publisher = None
                    pages = None
                    publication_year = None
                    edition_number = None
                    is_textbook = 'N'
                    
                    # Извлекаем ISBN, если он есть в заголовках
                    if "ISBN*" in header_indices:
                        isbn = row_values[header_indices["ISBN*"]]
                        if not isbn and is_isbn_required:
                            row_error_messages.append(f"Отсутствует обязательное поле 'ISBN*'")
                        elif isbn and is_isbn_required:
                            # В режиме работы по ISBN сначала проверяем, существует ли такая книга
                            cursor.execute("SELECT id, quantity FROM books WHERE isbn = ?", (isbn,))
                            existing_book = cursor.fetchone()
                            if existing_book:
                                # Книга найдена - устанавливаем book_id и текущее количество
                                book_id = existing_book[0]
                                current_quantity = existing_book[1]
                                logging.info(f"Строка {i}: Режим ISBN API: Найдена книга по ISBN {isbn}, ID: {book_id}")
                            else:
                                # Книга с таким ISBN не найдена, получаем данные через API
                                logging.info(f"Строка {i}: Режим ISBN API: Книга с ISBN {isbn} не найдена в БД, запрашиваем данные через API")
                                book_data = await fetch_book_by_isbn(isbn)
                                if book_data:
                                    title = book_data.get('title')
                                    author = book_data.get('author')
                                    theme = book_data.get('theme')
                                    description = book_data.get('description')
                                    publisher = book_data.get('publisher')
                                    pages = book_data.get('pages')
                                    publication_year = book_data.get('publication_year')
                                    logging.info(f"Строка {i}: Режим ISBN API: Получены данные о книге через API: {title} от {author}")
                                else:
                                    # Не удалось получить данные о книге через API
                                    row_error_messages.append(f"Не удалось получить данные о книге с ISBN {isbn} через API")
                                    logging.warning(f"Строка {i}: Режим ISBN API: Не удалось получить данные о книге с ISBN {isbn}")
                    elif "ISBN" in header_indices:
                        isbn = row_values[header_indices["ISBN"]]
                    
                    # Извлекаем остальные данные о книге из Excel, если они не были получены через API
                    if "Название книги*" in header_indices and not title:
                        title = row_values[header_indices["Название книги*"]]
                        if not title and is_title_required:  # Проверяем только если это обязательное поле
                            row_error_messages.append(f"Отсутствует обязательное поле 'Название книги*'")
                    elif "Название книги" in header_indices and not title:
                        title = row_values[header_indices["Название книги"]]
                        
                    if "Автор*" in header_indices and not author:
                        author = row_values[header_indices["Автор*"]]
                        if not author and is_title_required:  # Проверяем только если полный режим
                            row_error_messages.append(f"Отсутствует обязательное поле 'Автор*'")
                    elif "Автор" in header_indices and not author:
                        author = row_values[header_indices["Автор"]]
                        
                    if "Тематика(Жанр)*" in header_indices and not theme:
                        theme = row_values[header_indices["Тематика(Жанр)*"]]
                        if not theme and is_title_required:  # Проверяем только если полный режим
                            row_error_messages.append(f"Отсутствует обязательное поле 'Тематика(Жанр)*'")
                    elif "Тематика(Жанр)" in header_indices and not theme:
                        theme = row_values[header_indices["Тематика(Жанр)"]]
                    
                    if "Описание*" in header_indices and not description:
                        description = row_values[header_indices["Описание*"]]
                        if not description and is_title_required:  # Проверяем только если полный режим
                            row_error_messages.append(f"Отсутствует обязательное поле 'Описание*'")
                    elif "Описание" in header_indices and not description:
                        description = row_values[header_indices["Описание"]]
                    
                    if "Издательство*" in header_indices and not publisher:
                        publisher = row_values[header_indices["Издательство*"]]
                        if not publisher and is_title_required:  # Проверяем только если полный режим
                            row_error_messages.append(f"Отсутствует обязательное поле 'Издательство*'")
                    elif "Издательство" in header_indices and not publisher:
                        publisher = row_values[header_indices["Издательство"]]
                    
                    # Обработка числовых полей
                    try:
                        if "Кол-во страниц*" in header_indices and not pages:
                            pages_raw = row_values[header_indices["Кол-во страниц*"]]
                            if pages_raw is not None and pages_raw != "":
                                pages = int(pages_raw)
                            elif is_title_required:  # Проверяем только если полный режим
                                row_error_messages.append(f"Отсутствует обязательное поле 'Кол-во страниц*'")
                        elif "Кол-во страниц" in header_indices and not pages:
                            pages_raw = row_values[header_indices["Кол-во страниц"]]
                            if pages_raw is not None and pages_raw != "":
                                pages = int(pages_raw)
                    except (ValueError, TypeError):
                        if is_title_required:  # Проверяем только если полный режим
                            row_error_messages.append(f"Неверный формат числа в поле 'Кол-во страниц'")
                    
                    try:
                        if "Год издания*" in header_indices and not publication_year:
                            year_raw = row_values[header_indices["Год издания*"]]
                            if year_raw is not None and year_raw != "":
                                publication_year = int(year_raw)
                            elif is_title_required:  # Проверяем только если полный режим
                                row_error_messages.append(f"Отсутствует обязательное поле 'Год издания*'")
                        elif "Год издания" in header_indices and not publication_year:
                            year_raw = row_values[header_indices["Год издания"]]
                            if year_raw is not None and year_raw != "":
                                publication_year = int(year_raw)
                    except (ValueError, TypeError):
                        if is_title_required:  # Проверяем только если полный режим
                            row_error_messages.append(f"Неверный формат числа в поле 'Год издания'")
                    
                    # Номер издания (необязателен)
                    if "Номер издания (если есть)" in header_indices:
                        edition_number = row_values[header_indices["Номер издания (если есть)"]]
                    
                    # Учебник (необязательное поле для обоих режимов)
                    textbook_field = None
                    if "Учебник(Y или N)" in header_indices:
                        textbook_field = row_values[header_indices["Учебник(Y или N)"]]
                    
                    if textbook_field:
                        is_textbook = 'Y' if str(textbook_field).upper() == 'Y' else 'N'
                    
                    # --- Извлекаем данные для ЗАКУПКИ ---
                    purchase_date_raw = None
                    purchase_quantity_raw = None
                    price_per_unit_raw = None
                    supplier_raw = None
                    
                    if "Дата поставки*" in header_indices:
                        purchase_date_raw = row_values[header_indices["Дата поставки*"]]
                        if not purchase_date_raw:
                            row_error_messages.append(f"Отсутствует обязательное поле 'Дата поставки*'")
                    
                    if "Количество экземпляров*" in header_indices:
                        purchase_quantity_raw = row_values[header_indices["Количество экземпляров*"]]
                        if not purchase_quantity_raw:
                            row_error_messages.append(f"Отсутствует обязательное поле 'Количество экземпляров*'")
                    
                    if "Цена за экземпляр*" in header_indices:
                        price_per_unit_raw = row_values[header_indices["Цена за экземпляр*"]]
                        if not price_per_unit_raw:
                            row_error_messages.append(f"Отсутствует обязательное поле 'Цена за экземпляр*'")
                    
                    if "Поставщик*" in header_indices:
                        supplier_raw = row_values[header_indices["Поставщик*"]]
                        if not supplier_raw:
                            row_error_messages.append(f"Отсутствует обязательное поле 'Поставщик*'")
                            
                    # Проверяем, есть ли ошибки в данных строки
                    if row_error_messages:
                        # Для режима ISBN, если ISBN найден, мы можем быть более терпимыми к ошибкам
                        if is_isbn_required and isbn:
                            # Проверяем, существует ли книга с таким ISBN или получены данные через API
                            cursor.execute("SELECT id FROM books WHERE isbn = ?", (isbn,))
                            existing_book = cursor.fetchone()
                            if (existing_book or (title and author)) and purchase_quantity_raw:
                                # Если книга с таким ISBN существует или получены данные через API, и указано количество,
                                # можно продолжить, учитывая, что некоторые другие поля могут быть пустыми
                                logging.info(f"Строка {i}: Найдена книга по ISBN: {isbn} или получены данные через API. Продолжаем, несмотря на некоторые пустые поля")
                                # Очищаем ошибки для полей, которые можно заполнить автоматически
                                row_error_messages = [msg for msg in row_error_messages 
                                                    if not ("Дата поставки" in msg or "Цена за экземпляр" in msg or "Поставщик" in msg)]
                                # Устанавливаем значения по умолчанию для пустых полей
                                if not purchase_date_raw:
                                    purchase_date_raw = datetime.now()
                                    logging.info(f"Строка {i}: Установлена текущая дата для пустого поля 'Дата поставки*'")
                                if not price_per_unit_raw:
                                    price_per_unit_raw = 0.0
                                    logging.info(f"Строка {i}: Установлена цена 0 для пустого поля 'Цена за экземпляр*'")
                                if not supplier_raw:
                                    supplier_raw = "Неизвестный поставщик"
                                    logging.info(f"Строка {i}: Установлен поставщик 'Неизвестный поставщик' для пустого поля 'Поставщик*'")
                        
                        # Если все еще есть ошибки, пропускаем строку
                        if row_error_messages:
                            error_summary = f"Строка {i}: " + "; ".join(row_error_messages)
                            row_errors_details.append(error_summary)
                            logging.warning(error_summary)
                            error_rows.append(i)
                            skipped_rows += 1
                            continue
                    
                    # Если данные книги получены через API, но нет book_id, создаем новую книгу
                    if is_isbn_required and isbn and title and author and not book_id:
                        logging.info(f"Строка {i}: Создание новой книги по данным API: '{title}' автора '{author}'")
                        cursor.execute("""
                            INSERT INTO books (
                                title, author, theme, description, pages, edition_number,
                                publication_year, publisher, quantity, is_textbook, isbn
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """, (title, author, theme, description, pages, edition_number,
                             publication_year, publisher, 0, is_textbook, isbn))
                        book_id = cursor.lastrowid
                        current_quantity = 0
                        conn.commit()
                        logging.info(f"Строка {i}: Добавлена новая книга ID {book_id} '{title}' по данным API")
                        
                        try:
                            # Логируем действие администратора
                            log_admin_action(
                                admin_id=admin_id,
                                action_type="add_book",
                                book_id=book_id,
                                details=f"Добавлена книга через Excel по ISBN API: '{title}' (ISBN: {isbn})"
                            )
                        except Exception as log_error:
                            logging.warning(f"Ошибка при логировании действия администратора: {log_error}")

                    logging.info(f"Строка {i}: Данные книги - '{title}' автора '{author}'")
                    
                    supplier = supplier_raw  # Предполагаем, что поставщик - строка
                    
                    # Обработка числовых полей закупки
                    try: 
                        purchase_quantity = int(purchase_quantity_raw) if purchase_quantity_raw is not None else 0
                        if purchase_quantity <= 0:
                            row_error_messages.append(f"Количество экземпляров должно быть больше 0")
                    except (ValueError, TypeError): 
                        purchase_quantity = 0
                        row_error_messages.append(f"Неверный формат числа в поле 'Количество экземпляров*'")

                    try: 
                        price_per_unit = float(price_per_unit_raw) if price_per_unit_raw is not None else 0.0
                        if price_per_unit < 0:
                            row_error_messages.append(f"Цена не может быть отрицательной")
                    except (ValueError, TypeError): 
                        price_per_unit = 0.0
                        row_error_messages.append(f"Неверный формат числа в поле 'Цена за экземпляр*'")

                    # Обработка даты поставки
                    purchase_date_iso = None
                    if isinstance(purchase_date_raw, datetime):
                        purchase_date_iso = purchase_date_raw.strftime('%Y-%m-%d')
                    elif isinstance(purchase_date_raw, str):
                        try:
                            dt_obj = datetime.strptime(purchase_date_raw, '%d.%m.%Y')
                            purchase_date_iso = dt_obj.strftime('%Y-%m-%d')
                        except ValueError:
                            try:
                                dt_obj = datetime.strptime(purchase_date_raw, '%Y-%m-%d')
                                purchase_date_iso = dt_obj.strftime('%Y-%m-%d')
                            except ValueError:
                                row_error_messages.append(f"Неверный формат даты в поле 'Дата поставки*'. Используйте формат ДД.ММ.ГГГГ")
                                purchase_date_iso = None
                    else:
                        row_error_messages.append(f"Некорректные данные в поле 'Дата поставки*'")
                        purchase_date_iso = None
                    
                    # Проверяем еще раз, есть ли ошибки после обработки
                    if row_error_messages:
                        error_summary = f"Строка {i}: " + "; ".join(row_error_messages)
                        row_errors_details.append(error_summary)
                        logging.warning(error_summary)
                        error_rows.append(i)
                        skipped_rows += 1
                        continue
                    
                    logging.info(f"Строка {i}: Данные закупки - Дата: '{purchase_date_iso}', Кол-во: {purchase_quantity}, Цена: {price_per_unit}, Поставщик: '{supplier}'")
                    
                    # --- Поиск или создание книги ---
                    # Сначала ищем по ISBN, если он есть
                    if isbn:
                        cursor.execute("SELECT id, quantity FROM books WHERE isbn = ?", (isbn,))
                        existing_book = cursor.fetchone()
                        if existing_book:
                            book_id = existing_book[0]
                            current_quantity = existing_book[1]
                            # Обновляем данные, только если они предоставлены
                            update_fields = []
                            update_values = []
                            
                            if title:
                                update_fields.append("title = ?")
                                update_values.append(title)
                            if author:
                                update_fields.append("author = ?")
                                update_values.append(author)
                            if theme:
                                update_fields.append("theme = ?")
                                update_values.append(theme)
                            if description:
                                update_fields.append("description = ?")
                                update_values.append(description)
                            if pages:
                                update_fields.append("pages = ?")
                                update_values.append(pages)
                            if edition_number:
                                update_fields.append("edition_number = ?")
                                update_values.append(edition_number)
                            if publication_year:
                                update_fields.append("publication_year = ?")
                                update_values.append(publication_year)
                            if publisher:
                                update_fields.append("publisher = ?")
                                update_values.append(publisher)
                            
                            # Всегда обновляем количество и статус учебника
                            update_fields.append("quantity = ?")
                            update_values.append(current_quantity + purchase_quantity)
                            update_fields.append("is_textbook = ?")
                            update_values.append(is_textbook)
                            
                            if update_fields:
                                # Формируем и выполняем запрос на обновление
                                update_sql = f"UPDATE books SET {', '.join(update_fields)} WHERE id = ?"
                                update_values.append(book_id)
                                cursor.execute(update_sql, update_values)
                                logging.info(f"Строка {i}: Обновлена книга ID {book_id}")
                    
                    # Если не нашли по ISBN или его нет, ищем по названию и автору
                    if not book_id and title and author:
                        cursor.execute("SELECT id, quantity FROM books WHERE title = ? AND author = ?", (title, author))
                        existing_book = cursor.fetchone()

                        if existing_book:
                            book_id = existing_book[0]
                            current_quantity = existing_book[1]
                            # Обновляем существующую книгу
                            cursor.execute("""
                                UPDATE books SET
                                    theme = COALESCE(?, theme),
                                    description = COALESCE(?, description),
                                    pages = COALESCE(?, pages),
                                    edition_number = COALESCE(?, edition_number),
                                    publication_year = COALESCE(?, publication_year),
                                    publisher = COALESCE(?, publisher),
                                    quantity = ?,
                                    is_textbook = ?,
                                    isbn = COALESCE(?, isbn)
                                WHERE id = ?
                            """, (theme, description, pages, edition_number, publication_year, publisher,
                                 current_quantity + purchase_quantity, is_textbook, isbn, book_id))
                            logging.info(f"Строка {i}: Обновлена книга ID {book_id} по названию и автору")
                    
                    # Если книга не найдена ни по ISBN, ни по названию и автору, создаем новую
                    if not book_id:
                        # Проверяем режим работы и требования к данным
                        if is_isbn_required and isbn:
                            # В режиме ISBN книга должна уже существовать в базе данных
                            error_msg = f"Строка {i}: Книга с ISBN {isbn} не найдена в базе данных. В режиме закупок по ISBN книга должна уже существовать."
                            row_errors_details.append(error_msg)
                            logging.warning(error_msg)
                            error_rows.append(i)
                            skipped_rows += 1
                            continue
                        
                        # Для режима полных данных проверяем, что есть все необходимые поля для создания книги
                        if is_title_required:
                            if not title or not author:
                                error_msg = f"Строка {i}: Невозможно создать новую книгу без названия и автора"
                                row_errors_details.append(error_msg)
                                logging.warning(error_msg)
                                error_rows.append(i)
                                skipped_rows += 1
                                continue
                            
                            # Добавляем новую книгу
                            logging.info(f"Строка {i}: Создание новой книги с названием '{title}' и автором '{author}'")
                            cursor.execute("""
                                INSERT INTO books (
                                    title, author, theme, description, pages, edition_number,
                                    publication_year, publisher, quantity, is_textbook, isbn
                                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """, (title, author, theme, description, pages, edition_number,
                                publication_year, publisher, purchase_quantity, is_textbook, isbn))
                            book_id = cursor.lastrowid
                            logging.info(f"Строка {i}: Добавлена новая книга ID {book_id} '{title}'")
                            
                            # Делаем промежуточный коммит после добавления книги
                            conn.commit()
                            
                            try:
                                # Логируем действие администратора
                                log_admin_action(
                                    admin_id=admin_id,
                                    action_type="add_book",
                                    book_id=book_id,
                                    details=f"Добавлена книга через Excel: '{title}' (ISBN: {isbn or 'нет'})"
                                )
                            except Exception as log_error:
                                logging.warning(f"Ошибка при логировании действия администратора: {log_error}")
                        else:
                            # Режим ISBN, но книга не найдена
                            error_msg = f"Строка {i}: В режиме закупок по ISBN книга должна уже существовать в базе данных"
                            row_errors_details.append(error_msg)
                            logging.warning(error_msg)
                            error_rows.append(i)
                            skipped_rows += 1
                            continue
                    # Если всё в порядке и у нас есть book_id, добавляем закупку
                    if book_id and purchase_date_iso and purchase_quantity > 0:
                        logging.info(f"Строка {i}: Начинаем добавление закупки: ID книги={book_id}, дата={purchase_date_iso}, кол-во={purchase_quantity}, цена={price_per_unit}, поставщик={supplier}")
                        try:
                            # Добавляем запись о закупке
                            cursor.execute("""
                                INSERT INTO book_purchases (book_id, purchase_date, quantity, price_per_unit, supplier)
                                VALUES (?, ?, ?, ?, ?)
                            """, (book_id, purchase_date_iso, purchase_quantity, price_per_unit, supplier))
                            logging.info(f"Строка {i}: Закупка добавлена в таблицу book_purchases")
                            
                            # Делаем промежуточный коммит после добавления записи о закупке
                            conn.commit()
                                
                            # Создаем экземпляры книги
                            cursor.execute("SELECT COUNT(*) FROM book_copies WHERE book_id = ?", (book_id,))
                            current_copies = cursor.fetchone()[0]
                            logging.info(f"Строка {i}: Текущее количество экземпляров: {current_copies}")
                            
                            # Добавляем экземпляры небольшими партиями для предотвращения блокировок
                            batch_size = 10
                            for batch_start in range(0, purchase_quantity, batch_size):
                                batch_end = min(batch_start + batch_size, purchase_quantity)
                                for copy_num in range(batch_start, batch_end):
                                    copy_number = current_copies + copy_num + 1
                                    cursor.execute("""
                                        INSERT INTO book_copies (book_id, copy_number, status)
                                        VALUES (?, ?, 'available')
                                    """, (book_id, copy_number))
                                # Коммит после каждой партии
                                conn.commit()
                                
                            logging.info(f"Строка {i}: Добавлено {purchase_quantity} новых экземпляров книги")
                                    
                            added_purchases += 1
                            logging.info(f"Строка {i}: Добавлена закупка для книги ID {book_id}")
                            
                            try:
                                # Логируем действие администратора
                                log_admin_action(
                                    admin_id=admin_id,
                                    action_type="add_purchase",
                                    book_id=book_id,
                                    details=f"Добавлена закупка через Excel: {purchase_quantity} экз. книги '{title}' (ID: {book_id})"
                                )
                            except Exception as log_error:
                                logging.warning(f"Ошибка при логировании закупки: {log_error}")
                        except Exception as purchase_error:
                            error_msg = f"Строка {i}: Ошибка при добавлении закупки: {purchase_error}"
                            row_errors_details.append(error_msg)
                            logging.error(error_msg)
                            continue
                    else:
                        if not book_id:
                            logging.warning(f"Строка {i}: Не удалось добавить закупку - отсутствует book_id")
                        if not purchase_date_iso:
                            logging.warning(f"Строка {i}: Не удалось добавить закупку - отсутствует дата")
                        if not purchase_quantity or purchase_quantity <= 0:
                            logging.warning(f"Строка {i}: Не удалось добавить закупку - некорректное количество: {purchase_quantity}")
                    
                    processed_books += 1

                except Exception as row_error:
                    error_msg = f"Строка {i}: Общая ошибка обработки строки: {row_error}"
                    row_errors_details.append(error_msg)
                    logging.error(error_msg)
                    error_rows.append(i)
                    skipped_rows += 1
                    continue

            # Фиксируем транзакцию
            conn.commit()
            logging.info("Обработка Excel файла завершена.")

            # Формируем итоговое сообщение
            result_message = f"Файл успешно обработан. Добавлено/обновлено книг: {processed_books}. Добавлено закупок: {added_purchases}."
            
            if skipped_rows > 0:
                result_message += f"\n\nПРЕДУПРЕЖДЕНИЯ:\nПропущено строк с ошибками: {skipped_rows}"
                
                # Добавляем детали ошибок (максимум 15 сообщений)
                if row_errors_details:
                    max_errors = 15
                    errors_to_show = row_errors_details[:max_errors]
                    
                    if len(row_errors_details) > max_errors:
                        errors_to_show.append(f"... и еще {len(row_errors_details) - max_errors} ошибок")
                    
                    result_message += "\n\n" + "\n".join(errors_to_show)
            
            # Возвращаем результат
            return JSONResponse({"message": result_message})

    except Exception as e:
        logging.exception(f"Критическая ошибка при загрузке Excel файла: {e}")
        return JSONResponse({"error": f"Ошибка при обработке файла: {e}. Пожалуйста, убедитесь, что вы загружаете файл в правильном формате."}, status_code=500)

@router.get("/admin/users/template/download")
async def download_users_template(request: Request):
    if not is_admin(request):
        return RedirectResponse(url="/login")
    
    # Создаем Excel файл
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Пользователи"
    
    # Определяем заголовки
    headers = [
        "id_пользователя*", "Имя пользователя", "ФИО*", "Телефон*", "Роль*", "Класс"
    ]
    
    # Форматирование и заполнение шаблона
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = ws[f"{col_letter}1"]
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[col_letter].width = 20
    
    # Добавляем пример данных
    example_data = [
        "1234567890", "ivanov", "Иванов Иван Иванович", "+7 999 123 4567", "user", "9А"
    ]
    
    for col_num, value in enumerate(example_data, 1):
        ws[f"{get_column_letter(col_num)}2"] = value
    
    # Сохраняем в буфер
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Возвращаем файл
    return StreamingResponse(
        buffer, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=users_template.xlsx"}
    )

@router.post("/admin/users/upload")
async def upload_users_excel(request: Request):
    if not is_admin(request):
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    try:
        # Получаем файл из запроса
        form = await request.form()
        file = form.get("file")
        if not file:
            return JSONResponse({"error": "Файл не найден"}, status_code=400)
        
        # Читаем Excel файл
        contents = await file.read()
        wb = openpyxl.load_workbook(BytesIO(contents), data_only=True)
        ws = wb.active
        
        # Проверяем заголовки - используем первую строку напрямую
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = list(first_row)
        required_headers = ["id_пользователя*", "ФИО*", "Телефон*", "Роль*"]
        
        for header in required_headers:
            if header not in headers:
                return JSONResponse({"error": f"Отсутствует обязательный столбец: {header}"}, status_code=400)
        
        # Индексы столбцов
        id_idx = headers.index("id_пользователя*")
        username_idx = headers.index("Имя пользователя") if "Имя пользователя" in headers else None
        name_idx = headers.index("ФИО*")
        role_idx = headers.index("Роль*")
        class_idx = headers.index("Класс") if "Класс" in headers else None
        phone_idx = headers.index("Телефон*")
        
        with get_db() as conn:
            cursor = conn.cursor()
            added_count = 0
            
            # Начиная со второй строки (после заголовков), используем values_only для прямого получения значений
            for i, row_values in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), 2):
                if all(value is None for value in row_values):
                    continue  # Пропускаем пустые строки
                
                try:
                    # Извлекаем данные напрямую из значений
                    user_id = str(row_values[id_idx]) if row_values[id_idx] is not None else None
                    username = row_values[username_idx] if username_idx is not None else None
                    name = row_values[name_idx]
                    role = row_values[role_idx]
                    class_val = row_values[class_idx] if class_idx is not None else None
                    phone = row_values[phone_idx]
                    
                    # Проверяем обязательные поля
                    if not (user_id and name and role and phone):
                        continue
                        
                    # Проверяем правильность роли
                    if role not in ['admin', 'teacher', 'user']:
                        role = 'user'  # Установка роли по умолчанию
                    
                    # Проверяем существует ли пользователь
                    cursor.execute("SELECT id FROM users WHERE id = ?", (user_id,))
                    existing_user = cursor.fetchone()
                    
                    if existing_user:
                        # Обновляем существующего пользователя
                        cursor.execute("""
                            UPDATE users 
                            SET full_name = ?, role = ?, class = ?, phone = ?, username = ?
                            WHERE id = ?
                        """, (name, role, class_val, phone, username, user_id))
                    else:
                        # Добавляем нового пользователя
                        cursor.execute("""
                            INSERT INTO users (id, username, full_name, role, class, phone)
                            VALUES (?, ?, ?, ?, ?, ?)
                        """, (user_id, username, name, role, class_val, phone))
                    
                    # Фиксируем изменения
                    conn.commit()
                    added_count += 1
                    
                except Exception as e:
                    logging.error(f"Error processing row {i}: {e}")
            
            return JSONResponse({"success": True, "added": added_count})
        
    except Exception as e:
        logging.error(f"Error uploading Excel file: {e}")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.post("/users/{user_id}/edit")
async def edit_user(user_id: int, request: Request):
    try:
        # Проверка аутентификации и прав администратора
        # Здесь должна быть проверка прав доступа
        # В FastAPI это обычно делается через Depends
        
        # Получаем данные из формы
        form_data = await request.form()
        username = form_data.get('username')
        full_name = form_data.get('full_name')
        phone = form_data.get('phone')
        role = form_data.get('role')
        class_name = form_data.get('class')
        
        # Проверяем, существует ли пользователь
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT id FROM users WHERE id = ?", (user_id,))
            if not cursor.fetchone():
                return JSONResponse({"error": "Пользователь не найден"}, status_code=404)
            
            # Обновляем данные пользователя
            cursor.execute("""
                UPDATE users 
                SET username = ?, full_name = ?, phone = ?, role = ?, class = ?
                WHERE id = ?
            """, (username, full_name, phone, role, class_name, user_id))
            
            conn.commit()
            
            # Логируем действие
            log_admin_action(
                admin_id=1,  # Временно используем ID 1 для администратора
                action_type="edit_user",
                details=f"Отредактирован пользователь ID: {user_id}"
            )
            
            return JSONResponse({"success": True})
            
    except Exception as e:
        logging.error(f"Error editing user: {e}")
        return JSONResponse({"error": "Произошла ошибка при редактировании пользователя"}, status_code=500)

@router.get("/books/{book_id}/json")
async def get_book_json(request: Request, book_id: int):
    try:
        if not is_admin(request):
            raise HTTPException(status_code=403, detail="Доступ запрещен")
            
        with get_db() as conn:
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT * FROM books WHERE id = ?
            """, (book_id,))
            
            book_data = cursor.fetchone()
            
            if not book_data:
                raise HTTPException(status_code=404, detail="Книга не найдена")
            
            # Преобразуем в словарь
            book = dict(zip([column[0] for column in cursor.description], book_data))
            
            return book
    except Exception as e:
        logging.error(f"Error getting book JSON: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/overdue-books")
async def overdue_books_page(
    request: Request, 
    page: int = Query(1, ge=1), 
    search: str = Query(""), 
    class_filter: str = Query("", alias="class"),
    days_filter: Optional[str] = Query("", alias="days")
):
    """
    Страница со списком просроченных книг.
    Поддерживает фильтрацию по поисковому запросу, классу и количеству дней просрочки.
    """
    try:
        if not is_admin(request):
            return RedirectResponse(url="/login", status_code=302)
            
        with get_db() as conn:
            cursor = conn.cursor()
            
            # Параметры пагинации
            per_page = 15
            offset = (page - 1) * per_page
            
            # Базовый SQL запрос
            base_query = """
                SELECT
                    b.title,
                    b.author,
                    u.full_name as student_name,
                    u.class,
                    u.id as user_id,
                    bb.id as borrowed_id,
                    bb.borrow_date,
                    bb.return_date,
                    CAST(julianday(date('now')) - julianday(date(bb.return_date)) AS INTEGER) as days_overdue
                FROM
                    borrowed_books bb
                JOIN 
                    book_copies bc ON bb.copy_id = bc.id
                JOIN 
                    books b ON bc.book_id = b.id
                JOIN 
                    users u ON bb.user_id = u.id
                WHERE
                    bb.status = 'borrowed'
                    AND bb.return_date < date('now')
            """
            
            # Подготовка условий фильтрации
            conditions = []
            params = []
            
            if search:
                conditions.append("(b.title LIKE ? OR b.author LIKE ? OR u.full_name LIKE ?)")
                search_param = f"%{search}%"
                params.extend([search_param, search_param, search_param])
            
            if class_filter:
                conditions.append("u.class = ?")
                params.append(class_filter)
            
            # Проверяем, что days_filter - это число
            if days_filter and days_filter.isdigit():
                conditions.append("(julianday(date('now')) - julianday(date(bb.return_date))) >= ?")
                params.append(int(days_filter))
            
            # Добавляем условия фильтрации
            if conditions:
                base_query += " AND " + " AND ".join(conditions)
            
            # Добавляем сортировку по дням просрочки (убывание)
            base_query += " ORDER BY days_overdue DESC"
            
            # Запрос для подсчета общего количества
            count_query = f"""
                SELECT
                    COUNT(*)
                FROM
                    borrowed_books bb
                JOIN 
                    book_copies bc ON bb.copy_id = bc.id
                JOIN 
                    books b ON bc.book_id = b.id
                JOIN 
                    users u ON bb.user_id = u.id
                WHERE
                    bb.status = 'borrowed'
                    AND bb.return_date < date('now')
            """
            
            if conditions:
                count_query += " AND " + " AND ".join(conditions)
            
            # Выполняем запрос на получение общего количества
            cursor.execute(count_query, params)
            total_overdue = cursor.fetchone()[0]
            
            # Добавляем LIMIT и OFFSET для пагинации
            base_query += " LIMIT ? OFFSET ?"
            params.extend([per_page, offset])
            
            # Выполняем запрос на получение данных
            cursor.execute(base_query, params)
            overdue_data = cursor.fetchall()
            
            # Преобразуем в список словарей
            overdue_books = [dict(zip([column[0] for column in cursor.description], row)) for row in overdue_data]
            
            # Форматируем даты для отображения
            for book in overdue_books:
                book['borrow_date'] = datetime.fromisoformat(book['borrow_date']).strftime("%d.%m.%Y")
                book['return_date'] = datetime.fromisoformat(book['return_date']).strftime("%d.%m.%Y")
            
            # Получаем список классов для фильтра
            cursor.execute("SELECT DISTINCT class FROM users WHERE class IS NOT NULL AND class != '' ORDER BY class")
            classes = [row[0] for row in cursor.fetchall()]
            
            # Вычисляем общее количество страниц
            total_pages = (total_overdue + per_page - 1) // per_page if total_overdue > 0 else 1
            
            return templates.TemplateResponse("overdue_books.html", {
                "request": request,
                "overdue_books": overdue_books,
                "total_overdue": total_overdue,
                "current_page": page,
                "total_pages": total_pages,
                "search_query": search,
                "class_filter": class_filter,
                "days_filter": days_filter,
                "classes": classes,
                "max": max,
                "min": min
            })
            
    except Exception as e:
        logging.error(f"Error loading overdue books page: {e}")
        return templates.TemplateResponse("error.html", {"request": request, "error": f"Ошибка загрузки страницы просроченных книг: {e}"})

# Маршрут для отметки книги как возвращенной
@router.post("/borrowed-books/{borrowed_id}/return")
async def return_book(request: Request, borrowed_id: int):
    try:
        if not is_admin(request):
            raise HTTPException(status_code=403, detail="Доступ запрещен")
            
        with get_db() as conn:
            cursor = conn.cursor()
            
            # Проверяем существование записи о выдаче
            cursor.execute("SELECT copy_id FROM borrowed_books WHERE id = ? AND status = 'borrowed'", (borrowed_id,))
            result = cursor.fetchone()
            
            if not result:
                raise HTTPException(status_code=404, detail="Запись о выдаче не найдена или книга уже возвращена")
            
            # Обновляем статус книги
            cursor.execute("""
                UPDATE borrowed_books
                SET status = 'returned', actual_return_date = date('now')
                WHERE id = ?
            """, (borrowed_id,))
            
            conn.commit()
            
            return {"success": True, "message": "Книга успешно отмечена как возвращенная"}
            
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error returning book: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# Маршрут для продления срока возврата книги
@router.post("/borrowed-books/{borrowed_id}/extend")
async def extend_return_date(
    request: Request, 
    borrowed_id: int, 
    data: dict
):
    try:
        if not is_admin(request):
            raise HTTPException(status_code=403, detail="Доступ запрещен")
            
        new_return_date = data.get("new_return_date")
        if not new_return_date:
            raise HTTPException(status_code=400, detail="Не указана новая дата возврата")
        
        # Проверяем, что новая дата не раньше чем текущая
        today = datetime.now().date()
        try:
            parsed_date = datetime.strptime(new_return_date, "%Y-%m-%d").date()
            if parsed_date <= today:
                raise HTTPException(status_code=400, detail="Новая дата должна быть позже текущей даты")
        except ValueError:
            raise HTTPException(status_code=400, detail="Некорректный формат даты")
            
        with get_db() as conn:
            cursor = conn.cursor()
            
            # Проверяем существование записи о выдаче
            cursor.execute("SELECT id FROM borrowed_books WHERE id = ? AND status = 'borrowed'", (borrowed_id,))
            if not cursor.fetchone():
                raise HTTPException(status_code=404, detail="Запись о выдаче не найдена или книга уже возвращена")
            
            # Обновляем дату возврата
            cursor.execute("""
                UPDATE borrowed_books
                SET return_date = ?
                WHERE id = ?
            """, (new_return_date, borrowed_id))
            
            conn.commit()
            
            return {"success": True, "message": "Дата возврата успешно продлена"}
            
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error extending return date: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# Маршрут для отправки напоминания пользователю
@router.post("/users/{user_id}/send-reminder")
async def send_book_reminder(
    request: Request, 
    user_id: int, 
    data: dict
):
    try:
        if not is_admin(request):
            raise HTTPException(status_code=403, detail="Доступ запрещен")
            
        borrowed_id = data.get("borrowed_id")
        if not borrowed_id:
            raise HTTPException(status_code=400, detail="Не указан ID записи о выдаче")
            
        with get_db() as conn:
            cursor = conn.cursor()
            
            # Получаем информацию о книге и пользователе
            cursor.execute("""
                SELECT 
                    u.full_name, u.telegram_id, u.email,
                    b.title, b.author,
                    bb.return_date,
                    CAST(julianday(date('now')) - julianday(date(bb.return_date)) AS INTEGER) as days_overdue
                FROM 
                    borrowed_books bb
                JOIN 
                    book_copies bc ON bb.copy_id = bc.id
                JOIN 
                    books b ON bc.book_id = b.id
                JOIN 
                    users u ON bb.user_id = u.id
                WHERE 
                    bb.id = ? AND u.id = ?
            """, (borrowed_id, user_id))
            
            result = cursor.fetchone()
            if not result:
                raise HTTPException(status_code=404, detail="Информация о выдаче не найдена")
                
            user_info = dict(zip([column[0] for column in cursor.description], result))
            
            # Готовим текст напоминания
            reminder_text = f"""
                Уважаемый(ая) {user_info['full_name']}!
                
                Напоминаем, что вы должны были вернуть книгу "{user_info['title']}" автора {user_info['author']} 
                до {user_info['return_date']}. На данный момент просрочка составляет {user_info['days_overdue']} дней.
                
                Пожалуйста, верните книгу в библиотеку в ближайшее время.
                
                С уважением, библиотека школы.
            """
            
            # Сохраняем запись о напоминании
            cursor.execute("""
                INSERT INTO user_notifications (user_id, message, created_at, type, status)
                VALUES (?, ?, datetime('now'), 'overdue_reminder', 'pending')
            """, (user_id, reminder_text))
            
            notification_id = cursor.lastrowid
            
            conn.commit()
            
            # Здесь можно добавить логику отправки уведомления через Telegram или email
            # Это зависит от реализации вашего бота и системы уведомлений
            
            # Отмечаем напоминание как отправленное
            cursor.execute("""
                UPDATE user_notifications
                SET status = 'sent', sent_at = datetime('now')
                WHERE id = ?
            """, (notification_id,))
            
            conn.commit()
            
            return {"success": True, "message": "Напоминание успешно отправлено"}
            
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error sending reminder: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/reports")
async def reports_page(request: Request):
    """
    Страница отчетов библиотеки.
    В разработке.
    """
    try:
        if not is_admin(request):
            return RedirectResponse(url="/login", status_code=302)
            
        admin_info = get_admin_info(request)
        
        return templates.TemplateResponse("reports.html", {
            "request": request,
            "admin_info": admin_info
        })
            
    except Exception as e:
        logging.error(f"Error loading reports page: {e}")
        return templates.TemplateResponse("error.html", {
            "request": request, 
            "error": f"Ошибка загрузки страницы отчетов: {e}"
        })

@router.get("/api/books/isbn/{isbn}")
async def find_book_by_isbn(isbn: str):
    """
    API эндпоинт для поиска книги по ISBN.
    Возвращает данные о книге, если она найдена через Google Books API.
    """
    try:
        # Ищем книгу только через API
        book_data = await fetch_book_by_isbn(isbn)
        
        if book_data:
            # Успешно получили данные из API
            return JSONResponse({"book": book_data, "source": "api"})
        else:
            return JSONResponse(
                {"error": f"Книга с ISBN {isbn} не найдена через API"}, 
                status_code=404
            )
            
    except Exception as e:
        logging.error(f"Ошибка при поиске книги по ISBN {isbn}: {e}")
        return JSONResponse(
            {"error": f"Произошла ошибка при поиске книги: {str(e)}"}, 
            status_code=500
        )